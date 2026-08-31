from __future__ import annotations

import html
import json
import re
import unicodedata
import warnings
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable, Optional

try:
    from rapidfuzz import fuzz, process
except ImportError:  # Программа продолжит работать без приблизительного поиска.
    fuzz = None
    process = None


ProgressCallback = Callable[[int, str], None]
CancelCallback = Callable[[], bool]


def _load_workbook_quiet(path: str | Path, **kwargs):
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
            module=r"openpyxl\.styles\.stylesheet",
        )
        return load_workbook(path, **kwargs)


@dataclass
class DatabaseRecord:
    record_number: int
    source_file: str = ""
    source_record_number: int = 0
    isbns: list[str] = field(default_factory=list)
    titles: list[str] = field(default_factory=list)
    authors: list[str] = field(default_factory=list)
    primary_authors: list[str] = field(default_factory=list)
    organizations: list[str] = field(default_factory=list)
    inventory_numbers: list[str] = field(default_factory=list)
    publication: list[str] = field(default_factory=list)
    raw_record: str = ""

    @property
    def main_isbn(self) -> str:
        return self.isbns[0] if self.isbns else ""

    @property
    def main_title(self) -> str:
        return self.titles[0] if self.titles else ""

    @property
    def main_author(self) -> str:
        return self.authors[0] if self.authors else ""


@dataclass
class ExcelEntry:
    entry_id: int
    source_file: str
    sheet_name: str
    row_number: int
    author: str = ""
    title: str = ""
    isbn: str = ""
    registration_number: str = ""
    raw_data: dict[str, Any] = field(default_factory=dict)


@dataclass
class ForeignAgentEntry:
    entry_id: int
    source_file: str
    sheet_name: str
    row_number: int
    registry_number: str = ""
    name: str = ""
    participants: list[str] = field(default_factory=list)
    agent_type: str = ""
    inclusion_date: str = ""
    exclusion_date: str = ""
    raw_data: dict[str, Any] = field(default_factory=dict)

    @property
    def is_active(self) -> bool:
        return not bool(self.exclusion_date.strip())


@dataclass
class MatchResult:
    status: str
    method: str
    confidence: float
    excel: ExcelEntry
    database: Optional[DatabaseRecord] = None
    note: str = ""
    source_type: str = "Вещества"
    matched_value: str = ""
    foreign_agent: Optional[ForeignAgentEntry] = None


@dataclass
class ComparisonSummary:
    database_file: str
    excel_files: list[str]
    database_records: int
    database_records_with_isbn: int
    excel_rows: int
    matched_excel_rows: int
    unmatched_excel_rows: int
    result_rows: int
    exact_isbn_rows: int
    exact_title_rows: int
    probable_rows: int
    foreign_agents_file: str = ""
    foreign_agent_rows: int = 0
    matched_foreign_agent_rows: int = 0
    foreign_agent_result_rows: int = 0
    substance_matched_records: int = 0
    foreign_agent_matched_records: int = 0
    output_file: str = ""
    modified_database_file: str = ""
    modified_database_records: int = 0
    markers_already_present: int = 0
    markers_added: int = 0
    marker_duplicates_repaired: int = 0
    warnings: list[str] = field(default_factory=list)


@dataclass
class MarkerApplicationStats:
    already_present: int = 0
    added: int = 0
    duplicates_repaired: int = 0


HEADER_SYNONYMS = {
    "author": {
        "автор", "авторы", "фио автора", "author", "authors",
    },
    "title": {
        "заглавие", "название", "наименование", "название книги", "title", "book title",
    },
    "isbn": {
        "isbn", "isbn 10", "isbn 13", "исбн", "международный стандартный номер книги",
    },
    "registration_number": {
        "рег №", "рег номер", "регистрационный номер", "регистрационный №",
        "инв №", "инвентарный номер", "номер", "рег. №",
    },
}


FOREIGN_AGENT_HEADER_SYNONYMS = {
    "registry_number": {"№ п/п", "номер", "номер п/п"},
    "name": {
        "полное наименование прежнее наименование в случае его изменения фио псевдоним при наличии прежние фио в случае их изменения",
        "полное наименование фио псевдоним",
        "полное наименование фио",
        "наименование фио",
    },
    "participants": {"полное наименование или фио участников", "участники"},
    "agent_type": {"тип иностранного агента", "тип иноагента"},
    "inclusion_date": {
        "дата принятия минюстом россии решения о включении в реестр",
        "дата включения в реестр",
    },
    "exclusion_date": {
        "дата принятия минюстом россии решения об исключении из реестра при наличии",
        "дата исключения из реестра",
    },
}

SOURCE_SUBSTANCES = "Вещества"
SOURCE_FOREIGN_AGENTS = "Иностранные агенты"
DEFAULT_SUBSTANCE_MARKER = "^AIII"
DEFAULT_FOREIGN_AGENT_MARKER_TEMPLATE = "^AI^@{name}"
DEFAULT_AGE_MARKER = "^Z18+"
DEFAULT_SUBSTANCE_MARKER_FIELD = 333
DEFAULT_FOREIGN_AGENT_MARKER_FIELD = 333
DEFAULT_AGE_MARKER_FIELD = 900

TITLE_STOP_WORDS = {
    "роман", "романы", "повесть", "повести", "рассказ", "рассказы", "сборник",
    "издание", "изд", "учебник", "учебное", "пособие", "текст", "перевод",
    "английского", "англ", "русского", "рус", "книга", "кн", "том", "часть",
    "16", "18", "12", "6", "0",
}


class ComparisonCancelled(RuntimeError):
    pass


def _cancelled(cancel_cb: Optional[CancelCallback]) -> None:
    if cancel_cb and cancel_cb():
        raise ComparisonCancelled("Операция отменена пользователем")


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", safe_text(value)).lower().replace("ё", "е")
    text = re.sub(r"[\s._-]+", " ", text)
    text = re.sub(r"[^0-9a-zа-я №]+", "", text)
    return re.sub(r"\s+", " ", text).strip()


ISBN_PLACEHOLDERS = {"", "0", "-", "—", "–", "нет", "нет isbn", "без isbn", "б/н", "бн", "n/a", "na"}
ISBN_PATTERN = re.compile(
    r"(?<![0-9X])(?:97[89](?:[\s\-\u2010-\u2015]?[0-9]){10}|(?:[0-9](?:[\s\-\u2010-\u2015]?[0-9]){8}[\s\-\u2010-\u2015]?[0-9X]))(?![0-9X])",
    re.IGNORECASE,
)


def _isbn_checksum_valid(code: str) -> bool:
    if len(code) == 13 and code.isdigit():
        return sum((1 if index % 2 == 0 else 3) * int(char) for index, char in enumerate(code)) % 10 == 0
    if len(code) == 10 and code[:9].isdigit() and (code[9].isdigit() or code[9] == "X"):
        total = sum((10 - index) * int(char) for index, char in enumerate(code[:9]))
        total += 10 if code[9] == "X" else int(code[9])
        return total % 11 == 0
    return False


def extract_isbns(value: Any) -> list[str]:
    """Возвращает все корректные ISBN-10/ISBN-13 из одной ячейки."""
    text = unicodedata.normalize("NFKC", safe_text(value)).upper().replace("Х", "X").strip()
    if text.lower() in ISBN_PLACEHOLDERS:
        return []

    found: list[str] = []
    for match in ISBN_PATTERN.finditer(text):
        code = re.sub(r"[^0-9X]", "", match.group(0).upper())
        if _isbn_checksum_valid(code) and code not in found:
            found.append(code)

    # На случай ячейки, содержащей только ISBN с необычной пунктуацией.
    if not found:
        compact = re.sub(r"[^0-9X]", "", text)
        if _isbn_checksum_valid(compact):
            found.append(compact)
    return found


def normalize_isbn(value: Any) -> str:
    candidates = extract_isbns(value)
    return candidates[0] if candidates else ""


def normalize_title(value: Any) -> str:
    text = unicodedata.normalize("NFKC", safe_text(value)).lower().replace("ё", "е")
    text = re.sub(r"\[[^\]]*\]|\([^)]*\)", " ", text)
    text = re.sub(r"\b\d+\+\b", " ", text)
    text = re.sub(r"[^0-9a-zа-я]+", " ", text)
    tokens = [token for token in text.split() if token not in TITLE_STOP_WORDS]
    return " ".join(tokens)


def normalize_author(value: Any) -> str:
    text = unicodedata.normalize("NFKC", safe_text(value)).lower().replace("ё", "е")
    text = re.sub(r"[^a-zа-я]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _author_identity(value: Any) -> tuple[str, tuple[str, ...]] | None:
    raw = unicodedata.normalize("NFKC", safe_text(value)).strip()
    if not raw:
        return None
    # Редакторские пометы вроде «(гл. ред.)» не являются частью ФИО.
    raw = re.sub(r"\([^)]*\)", " ", raw)
    first_person = re.split(r"[;\n]", raw, maxsplit=1)[0].strip()
    tokens = normalize_author(first_person).split()
    if not tokens:
        return None

    leading_initials: list[str] = []
    index = 0
    while index < len(tokens) - 1 and len(tokens[index]) == 1:
        leading_initials.append(tokens[index])
        index += 1

    if leading_initials and index < len(tokens):
        # «Л. Н. Толстой» / «Р. Фасхутдинов».
        return tokens[index], tuple(leading_initials[:2])

    # «Толстой Л. Н.», «Толстой Лев Николаевич», «Ильина, В. В.».
    surname = tokens[0]
    initials = tuple(token[0] for token in tokens[1:3] if token)
    return surname, initials


def author_surname(value: Any) -> str:
    identity = _author_identity(value)
    return identity[0] if identity else ""


def _extract_subfield(value: str, code: str) -> str:
    match = re.search(rf"\^{re.escape(code)}([^\^]*)", value)
    return match.group(1).strip() if match else ""


def _read_text_file_with_encoding(path: str | Path) -> tuple[str, str]:
    """Читает файл без изменения исходных переносов строк и BOM."""
    path = Path(path)
    raw = path.read_bytes()
    if raw.startswith(b"\xef\xbb\xbf"):
        return raw.decode("utf-8-sig"), "utf-8-sig"

    last_error: Optional[Exception] = None
    for encoding in ("utf-8", "cp1251"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"Не удалось определить кодировку файла {path.name}: {last_error}")


def _read_text_file(path: str | Path) -> str:
    text, _ = _read_text_file_with_encoding(path)
    return text


def database_record_from_tag_values(
    record_number: int,
    tag_values: Iterable[tuple[int | str, str]],
    *,
    source_file: str = "",
    source_record_number: int | None = None,
    raw_record: str = "",
) -> DatabaseRecord:
    """Создаёт DatabaseRecord из полей ИРБИС без промежуточного TXT-файла."""
    fields: dict[str, list[str]] = defaultdict(list)
    for tag, value in tag_values:
        try:
            key = str(int(tag))
        except (TypeError, ValueError):
            key = str(tag).strip().lstrip("0") or "0"
        fields[key].append(safe_text(value))

    isbns = [
        value for value in (_extract_subfield(item, "A") for item in fields.get("10", [])) if value
    ]
    titles = [
        value for value in (_extract_subfield(item, "A") for item in fields.get("200", [])) if value
    ]

    authors: list[str] = []
    primary_authors: list[str] = []
    for tag in ("700", "701", "702"):
        for item in fields.get(tag, []):
            parts = [
                _extract_subfield(item, "A"),
                _extract_subfield(item, "B"),
                _extract_subfield(item, "G"),
            ]
            author = " ".join(part for part in parts if part).strip()
            if author:
                authors.append(author)
                if tag in {"700", "701"}:
                    primary_authors.append(author)

    organizations: list[str] = []
    for tag in ("710", "711", "712"):
        for item in fields.get(tag, []):
            organization = _extract_subfield(item, "A")
            if organization:
                organizations.append(organization)

    inventory_numbers = [
        value for value in (_extract_subfield(item, "B") for item in fields.get("910", [])) if value
    ]

    return DatabaseRecord(
        record_number=record_number,
        source_file=source_file,
        source_record_number=source_record_number if source_record_number is not None else record_number,
        isbns=isbns,
        titles=titles,
        authors=authors,
        primary_authors=primary_authors,
        organizations=organizations,
        inventory_numbers=inventory_numbers,
        publication=fields.get("210", []),
        raw_record=raw_record,
    )


def parse_database(
    path: str | Path,
    progress_cb: Optional[ProgressCallback] = None,
    cancel_cb: Optional[CancelCallback] = None,
) -> list[DatabaseRecord]:
    path = Path(path)
    if progress_cb:
        progress_cb(2, f"Чтение текстовой базы: {path.name}")
    text = _read_text_file(path)
    raw_records = [part for part in re.split(r"\r?\n\*{5}\s*(?:\r?\n|$)", text) if part.strip()]
    records: list[DatabaseRecord] = []
    total = max(len(raw_records), 1)

    for index, raw_record in enumerate(raw_records, start=1):
        if index % 250 == 0:
            _cancelled(cancel_cb)
            if progress_cb:
                progress_cb(2 + int(index / total * 23), f"Разбор базы: {index:,} из {total:,}")

        tag_values: list[tuple[int, str]] = []
        for line in raw_record.splitlines():
            match = re.match(r"#(\d+):\s?(.*)$", line.strip())
            if match:
                tag_values.append((int(match.group(1)), match.group(2)))

        records.append(
            database_record_from_tag_values(
                index,
                tag_values,
                source_file=str(path),
                source_record_number=index,
                raw_record=raw_record,
            )
        )

    if progress_cb:
        progress_cb(25, f"База загружена: {len(records):,} записей")
    return records


def _detect_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int], list[str]]:
    best_score = 0
    best_row = -1
    best_map: dict[str, int] = {}
    best_headers: list[str] = []

    for row_index, row in enumerate(rows[:100]):
        mapping: dict[str, int] = {}
        headers = [safe_text(value) or f"Столбец {column + 1}" for column, value in enumerate(row)]
        for column, value in enumerate(row):
            normalized = normalize_header(value)
            for key, synonyms in HEADER_SYNONYMS.items():
                normalized_synonyms = {normalize_header(item) for item in synonyms}
                if normalized in normalized_synonyms and key not in mapping:
                    mapping[key] = column
                    break
        score = len(mapping)
        if score > best_score:
            best_score = score
            best_row = row_index
            best_map = mapping
            best_headers = headers

    if best_score == 0:
        raise ValueError(
            "Не удалось найти строку заголовков. Нужен хотя бы один столбец: ISBN, Заглавие/Название или Автор."
        )
    return best_row, best_map, best_headers


def _make_entry(
    entry_id: int,
    source_file: Path,
    sheet_name: str,
    row_number: int,
    row: Iterable[Any],
    mapping: dict[str, int],
    headers: list[str],
) -> ExcelEntry:
    values = list(row)

    def get(key: str) -> str:
        column = mapping.get(key)
        if column is None or column >= len(values):
            return ""
        return safe_text(values[column])

    raw_data: dict[str, Any] = {}
    for index, value in enumerate(values):
        if value is None or safe_text(value) == "":
            continue
        header = headers[index] if index < len(headers) else f"Столбец {index + 1}"
        raw_data[header] = safe_text(value)

    raw_isbn = get("isbn")
    isbn = "" if raw_isbn.strip().lower() in ISBN_PLACEHOLDERS else raw_isbn

    return ExcelEntry(
        entry_id=entry_id,
        source_file=str(source_file),
        sheet_name=sheet_name,
        row_number=row_number,
        author=get("author"),
        title=get("title"),
        isbn=isbn,
        registration_number=get("registration_number"),
        raw_data=raw_data,
    )


def _read_xlsx_entries(
    path: Path,
    start_entry_id: int,
    cancel_cb: Optional[CancelCallback] = None,
) -> tuple[list[ExcelEntry], list[str]]:
    entries: list[ExcelEntry] = []
    warnings: list[str] = []
    workbook = _load_workbook_quiet(path, read_only=True, data_only=True)
    entry_id = start_entry_id

    try:
        for worksheet in workbook.worksheets:
            _cancelled(cancel_cb)
            if worksheet.max_row == 1 and worksheet.max_column == 1:
                worksheet.reset_dimensions()
            max_preview_row = 100 if worksheet.max_row is None else min(100, worksheet.max_row)
            preview = list(worksheet.iter_rows(min_row=1, max_row=max_preview_row, values_only=True))
            if not preview or not any(any(value is not None for value in row) for row in preview):
                continue
            try:
                header_index, mapping, headers = _detect_header(preview)
            except ValueError as exc:
                warnings.append(f"{path.name}, лист «{worksheet.title}»: {exc}")
                continue

            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=header_index + 2, values_only=True), start=header_index + 2
            ):
                if row_number % 500 == 0:
                    _cancelled(cancel_cb)
                entry = _make_entry(entry_id, path, worksheet.title, row_number, row, mapping, headers)
                if entry.isbn or entry.title or entry.author:
                    entries.append(entry)
                    entry_id += 1
    finally:
        workbook.close()

    return entries, warnings


def _read_xls_entries(
    path: Path,
    start_entry_id: int,
    cancel_cb: Optional[CancelCallback] = None,
) -> tuple[list[ExcelEntry], list[str]]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("Для файлов .xls требуется пакет xlrd. Запустите start.bat ещё раз.") from exc

    entries: list[ExcelEntry] = []
    warnings: list[str] = []
    workbook = xlrd.open_workbook(path)
    entry_id = start_entry_id

    for sheet in workbook.sheets():
        _cancelled(cancel_cb)
        preview = [tuple(sheet.row_values(index)) for index in range(min(100, sheet.nrows))]
        if not preview or not any(any(safe_text(value) for value in row) for row in preview):
            continue
        try:
            header_index, mapping, headers = _detect_header(preview)
        except ValueError as exc:
            warnings.append(f"{path.name}, лист «{sheet.name}»: {exc}")
            continue

        for row_index in range(header_index + 1, sheet.nrows):
            if row_index % 500 == 0:
                _cancelled(cancel_cb)
            row = tuple(sheet.row_values(row_index))
            entry = _make_entry(entry_id, path, sheet.name, row_index + 1, row, mapping, headers)
            if entry.isbn or entry.title or entry.author:
                entries.append(entry)
                entry_id += 1

    return entries, warnings


def _deduplicate_cross_sheet_entries(entries: list[ExcelEntry]) -> tuple[list[ExcelEntry], int]:
    """Убирает зеркальные копии одной записи на разных листах, не трогая повторы внутри листа."""
    first_sheet_by_key: dict[tuple[Any, ...], str] = {}
    result: list[ExcelEntry] = []
    skipped = 0

    for entry in entries:
        key = (
            normalize_author(entry.author),
            normalize_title(entry.title),
            tuple(extract_isbns(entry.isbn)),
            normalize_header(entry.registration_number),
        )
        first_sheet = first_sheet_by_key.get(key)
        if first_sheet is None:
            first_sheet_by_key[key] = entry.sheet_name
            result.append(entry)
        elif first_sheet == entry.sheet_name:
            # Повтор на одном и том же листе может означать разные экземпляры книги.
            result.append(entry)
        else:
            skipped += 1

    return result, skipped


def read_excel_entries(
    paths: list[str | Path],
    progress_cb: Optional[ProgressCallback] = None,
    cancel_cb: Optional[CancelCallback] = None,
) -> tuple[list[ExcelEntry], list[str]]:
    all_entries: list[ExcelEntry] = []
    warnings: list[str] = []
    total = max(len(paths), 1)

    for file_index, source_path in enumerate(paths, start=1):
        _cancelled(cancel_cb)
        path = Path(source_path)
        if progress_cb:
            progress_cb(
                27 + int((file_index - 1) / total * 18),
                f"Чтение Excel {file_index} из {total}: {path.name}",
            )

        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            entries, file_warnings = _read_xlsx_entries(path, len(all_entries) + 1, cancel_cb)
        elif suffix == ".xls":
            entries, file_warnings = _read_xls_entries(path, len(all_entries) + 1, cancel_cb)
        else:
            warnings.append(f"Файл {path.name} пропущен: неподдерживаемое расширение {suffix}")
            continue

        entries, duplicate_count = _deduplicate_cross_sheet_entries(entries)
        if duplicate_count:
            file_warnings.append(
                f"{path.name}: исключено повторов одних и тех же записей на других листах: {duplicate_count}."
            )

        for entry in entries:
            entry.entry_id = len(all_entries) + 1
            all_entries.append(entry)
        warnings.extend(file_warnings)

    if progress_cb:
        progress_cb(45, f"Excel-строк для проверки: {len(all_entries):,}")
    return all_entries, warnings



def _detect_foreign_agent_header(
    rows: list[tuple[Any, ...]],
) -> tuple[int, dict[str, int], list[str]]:
    normalized_synonyms = {
        key: {normalize_header(item) for item in values}
        for key, values in FOREIGN_AGENT_HEADER_SYNONYMS.items()
    }
    best_score = 0
    best_row = -1
    best_map: dict[str, int] = {}
    best_headers: list[str] = []

    for row_index, row in enumerate(rows[:100]):
        mapping: dict[str, int] = {}
        headers = [safe_text(value) or f"Столбец {column + 1}" for column, value in enumerate(row)]
        for column, value in enumerate(row):
            normalized = normalize_header(value)
            for key, synonyms in normalized_synonyms.items():
                if normalized in synonyms and key not in mapping:
                    mapping[key] = column
                    break
        score = len(mapping)
        if "name" in mapping:
            score += 5
        if score > best_score:
            best_score = score
            best_row = row_index
            best_map = mapping
            best_headers = headers

    if best_row < 0 or "name" not in best_map:
        raise ValueError(
            "Не удалось найти столбец с полным наименованием/ФИО иностранного агента."
        )
    return best_row, best_map, best_headers


def _split_registry_participants(value: Any) -> list[str]:
    text = safe_text(value).strip()
    if not text:
        return []
    if text.startswith("[") and text.endswith("]"):
        text = text[1:-1].strip()
    parts = re.split(r'\s*,\s*(?=(?:[^\"]*\"[^\"]*\")*[^\"]*$)', text)
    return [part.strip() for part in parts if part.strip()]


def _read_foreign_agents_xlsx(
    path: Path,
    cancel_cb: Optional[CancelCallback] = None,
) -> tuple[list[ForeignAgentEntry], list[str]]:
    entries: list[ForeignAgentEntry] = []
    warnings: list[str] = []
    workbook = _load_workbook_quiet(path, read_only=True, data_only=True)
    entry_id = 1

    try:
        for worksheet in workbook.worksheets:
            _cancelled(cancel_cb)
            if worksheet.max_row == 1 and worksheet.max_column == 1:
                worksheet.reset_dimensions()
            max_preview_row = 100 if worksheet.max_row is None else min(100, worksheet.max_row)
            preview = list(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=max_preview_row,
                    values_only=True,
                )
            )
            if not preview or not any(any(value is not None for value in row) for row in preview):
                continue
            try:
                header_index, mapping, headers = _detect_foreign_agent_header(preview)
            except ValueError as exc:
                warnings.append(f"{path.name}, лист «{worksheet.title}»: {exc}")
                continue

            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=header_index + 2, values_only=True),
                start=header_index + 2,
            ):
                if row_number % 500 == 0:
                    _cancelled(cancel_cb)
                values = list(row)

                def get(key: str) -> str:
                    column = mapping.get(key)
                    if column is None or column >= len(values):
                        return ""
                    return safe_text(values[column])

                name = get("name")
                if not name:
                    continue

                raw_data: dict[str, Any] = {}
                for index, value in enumerate(values):
                    if value is None or safe_text(value) == "":
                        continue
                    header = headers[index] if index < len(headers) else f"Столбец {index + 1}"
                    raw_data[header] = safe_text(value)

                entry = ForeignAgentEntry(
                    entry_id=entry_id,
                    source_file=str(path),
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    registry_number=get("registry_number"),
                    name=name,
                    participants=_split_registry_participants(get("participants")),
                    agent_type=get("agent_type"),
                    inclusion_date=get("inclusion_date"),
                    exclusion_date=get("exclusion_date"),
                    raw_data=raw_data,
                )
                # Для проверки используются только действующие записи. Исключённые остаются
                # в исходном файле, но не должны приводить к новым меткам в библиотечной базе.
                if entry.is_active:
                    entries.append(entry)
                    entry_id += 1
    finally:
        workbook.close()

    return entries, warnings


def read_foreign_agent_entries(
    path: str | Path | None,
    progress_cb: Optional[ProgressCallback] = None,
    cancel_cb: Optional[CancelCallback] = None,
) -> tuple[list[ForeignAgentEntry], list[str]]:
    if not path:
        return [], []
    source = Path(path)
    if progress_cb:
        progress_cb(47, f"Чтение реестра иностранных агентов: {source.name}")
    if source.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError("Реестр иностранных агентов должен быть файлом Excel формата .xlsx или .xlsm.")
    entries, warnings = _read_foreign_agents_xlsx(source, cancel_cb)
    if progress_cb:
        progress_cb(52, f"Действующих записей в реестре иностранных агентов: {len(entries):,}")
    return entries, warnings


def _registry_plain_name(value: str) -> str:
    text = unicodedata.normalize("NFKC", safe_text(value)).replace("ё", "е").replace("Ё", "Е")
    text = re.sub(r'[«»„“”"]', ' ', text)
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[^0-9A-Za-zА-Яа-я]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _name_order_variants(value: str) -> list[str]:
    clean = _registry_plain_name(value)
    tokens = clean.split()
    variants = [clean] if clean else []
    if len(tokens) == 2:
        variants.append(f"{tokens[1]} {tokens[0]}")
    return list(dict.fromkeys(item for item in variants if item))


def _registry_name_variants(value: str, *, is_person: bool) -> list[str]:
    raw = safe_text(value)
    variants: list[str] = []
    quoted_values = re.findall(r'[«\"]([^»\"]+)[»\"]', raw)
    pseudonym_values = re.findall(r"\(\s*псевдоним\s*:\s*([^)]+)\)", raw, flags=re.IGNORECASE)

    if is_person:
        # Для ФИО кавычки обычно содержат псевдоним или прежнее имя.
        main_value = _registry_plain_name(re.sub(r'[«\"]([^»\"]+)[»\"]', ' ', raw))
        main_value = _registry_plain_name(re.sub(r"\(\s*псевдоним\s*:\s*[^)]+\)", " ", main_value, flags=re.IGNORECASE))
    else:
        # Для организаций сохраняем полное официальное наименование. Не добавляем
        # отдельно общую юридическую форму вроде «общество с ограниченной ответственностью».
        main_value = _registry_plain_name(raw)
    if main_value:
        variants.append(main_value)

    for quoted in quoted_values:
        for part in re.split(r"\s*,\s*", quoted):
            for clean in _name_order_variants(part):
                if clean and len(clean) >= 4:
                    variants.append(clean)
    for pseudonym in pseudonym_values:
        for part in re.split(r"\s*,\s*", pseudonym):
            for clean in _name_order_variants(part):
                if clean and len(clean) >= 4:
                    variants.append(clean)
    return list(dict.fromkeys(variants))


def _person_identity(value: str) -> tuple[str, tuple[str, ...], tuple[str, ...]] | None:
    tokens = re.findall(r"[A-Za-zА-Яа-я]+", _registry_plain_name(value).lower())
    if len(tokens) < 2:
        return None
    surname = tokens[0]
    full_names = tuple(token for token in tokens[1:] if len(token) > 1)
    if full_names:
        initials = tuple(token[0] for token in full_names[:2])
    else:
        initials = tuple(token[0] for token in tokens[1:3] if token)
    if not surname or not initials:
        return None
    return surname, full_names, initials


def _person_identity_match_kind(
    database_identity: tuple[str, tuple[str, ...], tuple[str, ...]],
    registry_identity: tuple[str, tuple[str, ...], tuple[str, ...]],
) -> str | None:
    """Возвращает тип подтверждения ФИО или None, если личности не совпадают.

    single_initial — ослабленное, но допустимое совпадение: фамилия + один
    инициал из библиотечной записи должны точно совпасть с фамилией и первым
    именем из реестра. Такой случай выводится в отчёт с предупреждением.
    """
    db_surname, db_full_names, db_initials = database_identity
    reg_surname, reg_full_names, reg_initials = registry_identity
    if db_surname != reg_surname:
        return None

    if db_full_names:
        # Полные имена из библиографической записи имеют приоритет над инициалами:
        # это исключает совпадения вроде «Петров Александр» и «Петров Алексей».
        if len(reg_full_names) < len(db_full_names):
            return None
        compare_count = min(len(db_full_names), 2)
        return "full_name" if db_full_names[:compare_count] == reg_full_names[:compare_count] else None

    if len(db_initials) >= 2:
        return "two_initials" if db_initials[:2] == reg_initials[:2] else None

    if len(db_initials) == 1 and reg_initials:
        return "single_initial" if db_initials[0] == reg_initials[0] else None

    return None


def _person_identities_match(
    database_identity: tuple[str, tuple[str, ...], tuple[str, ...]],
    registry_identity: tuple[str, tuple[str, ...], tuple[str, ...]],
) -> bool:
    return _person_identity_match_kind(database_identity, registry_identity) is not None


@dataclass(frozen=True)
class ForeignAgentSearchTerm:
    entry_id: int
    value: str
    normalized: str
    kind: str
    is_person: bool
    person_identity: tuple[str, tuple[str, ...], tuple[str, ...]] | None


class ForeignAgentIndex:
    def __init__(self, entries: list[ForeignAgentEntry]) -> None:
        self.entries = {entry.entry_id: entry for entry in entries}
        self.by_exact: dict[str, list[ForeignAgentSearchTerm]] = defaultdict(list)
        self.by_person_surname: dict[str, list[ForeignAgentSearchTerm]] = defaultdict(list)

        for entry in entries:
            terms: list[tuple[str, str, bool]] = []
            main_is_person = "физичес" in normalize_author(entry.agent_type)
            terms.extend(
                (variant, "ФИО/наименование", main_is_person)
                for variant in _registry_name_variants(entry.name, is_person=main_is_person)
            )
            for participant in entry.participants:
                terms.extend(
                    (variant, "Участник", True)
                    for variant in _registry_name_variants(participant, is_person=True)
                )

            seen: set[tuple[str, str, bool]] = set()
            for value, kind, is_person in terms:
                normalized = normalize_author(value)
                if not normalized or (normalized, kind, is_person) in seen:
                    continue
                seen.add((normalized, kind, is_person))
                identity = _person_identity(value) if is_person else None
                term = ForeignAgentSearchTerm(
                    entry_id=entry.entry_id,
                    value=value,
                    normalized=normalized,
                    kind=kind,
                    is_person=is_person,
                    person_identity=identity,
                )
                self.by_exact[normalized].append(term)
                if identity is not None:
                    self.by_person_surname[identity[0]].append(term)

    def match_record(
        self, record: DatabaseRecord
    ) -> list[tuple[ForeignAgentEntry, ForeignAgentSearchTerm, str, str, float]]:
        matched: dict[
            tuple[int, str, str],
            tuple[ForeignAgentEntry, ForeignAgentSearchTerm, str, str, float],
        ] = {}

        # Проверяем все персональные поля ответственности: #700, #701 и #702.
        # В parse_database они уже собраны в record.authors.
        for author in record.authors:
            author_normalized = normalize_author(author)
            database_identity = _person_identity(author)
            candidates: list[tuple[ForeignAgentSearchTerm, str | None]] = [
                (term, "exact") for term in self.by_exact.get(author_normalized, [])
            ]
            if database_identity is not None:
                for term in self.by_person_surname.get(database_identity[0], []):
                    if term.person_identity is None:
                        continue
                    match_kind = _person_identity_match_kind(database_identity, term.person_identity)
                    if match_kind is not None:
                        candidates.append((term, match_kind))

            for term, match_kind in candidates:
                if not term.is_person:
                    continue
                entry = self.entries[term.entry_id]
                note = term.kind
                confidence = 100.0
                if match_kind == "single_initial":
                    note = f"{term.kind}; фамилия + один инициал — проверить вручную"
                    confidence = 90.0
                key = (entry.entry_id, term.normalized, "Автор")
                previous = matched.get(key)
                candidate = (entry, term, "Автор", note, confidence)
                if previous is None or confidence > previous[4]:
                    matched[key] = candidate

        for organization in record.organizations:
            organization_normalized = normalize_author(organization)
            for term in self.by_exact.get(organization_normalized, []):
                if term.is_person:
                    continue
                entry = self.entries[term.entry_id]
                key = (entry.entry_id, term.normalized, "Организация")
                matched[key] = (entry, term, "Организация", term.kind, 100.0)

        for title in record.titles:
            title_normalized = normalize_author(title)
            for term in self.by_exact.get(title_normalized, []):
                # Названия книг сравниваются только с наименованиями организаций/проектов,
                # а не с ФИО физических лиц или участников.
                entry = self.entries[term.entry_id]
                if term.is_person or term.kind == "Участник":
                    continue
                if len(term.normalized.split()) < 2:
                    continue
                key = (entry.entry_id, term.normalized, "Название")
                matched[key] = (entry, term, "Название", term.kind, 100.0)

        return list(matched.values())


def compare_foreign_agents(
    records: list[DatabaseRecord],
    entries: list[ForeignAgentEntry],
    progress_cb: Optional[ProgressCallback] = None,
    cancel_cb: Optional[CancelCallback] = None,
) -> list[MatchResult]:
    if not entries:
        return []
    index = ForeignAgentIndex(entries)
    results: list[MatchResult] = []
    total = max(len(records), 1)

    for position, record in enumerate(records, start=1):
        if position % 250 == 0:
            _cancelled(cancel_cb)
            if progress_cb:
                progress_cb(72 + int(position / total * 8), f"Сверка с иноагентами: {position:,} из {total:,}")
        for entry, term, database_field, match_note, confidence in index.match_record(record):
            synthetic_entry = ExcelEntry(
                entry_id=entry.entry_id,
                source_file=entry.source_file,
                sheet_name=entry.sheet_name,
                row_number=entry.row_number,
                author=entry.name,
                title=term.value,
                registration_number=entry.registry_number,
                raw_data=entry.raw_data,
            )
            results.append(
                MatchResult(
                    status="Совпадение",
                    method=f"Реестр иностранных агентов: {database_field}",
                    confidence=confidence,
                    excel=synthetic_entry,
                    database=record,
                    note=match_note,
                    source_type=SOURCE_FOREIGN_AGENTS,
                    matched_value=term.value,
                    foreign_agent=entry,
                )
            )
    return results


def compare_substance_entries(
    index: "DatabaseIndex",
    entries: list[ExcelEntry],
    use_isbn_matching: bool,
    use_title_fallback: bool,
    use_fuzzy: bool,
    fuzzy_threshold: int,
    progress_cb: Optional[ProgressCallback] = None,
    cancel_cb: Optional[CancelCallback] = None,
) -> tuple[list[MatchResult], set[int], set[int], set[int], set[int]]:
    results: list[MatchResult] = []
    matched_entry_ids: set[int] = set()
    exact_isbn_ids: set[int] = set()
    exact_title_ids: set[int] = set()
    probable_ids: set[int] = set()
    total = max(len(entries), 1)

    for position, entry in enumerate(entries, start=1):
        if position % 100 == 0:
            _cancelled(cancel_cb)
            if progress_cb:
                progress_cb(53 + int(position / total * 18), f"Сравнение по веществам: {position:,} из {total:,}")

        entry_results = index.match(
            entry,
            use_isbn_matching,
            use_title_fallback,
            use_fuzzy,
            fuzzy_threshold,
        )
        results.extend(entry_results)
        if any(item.status == "Совпадение" for item in entry_results):
            matched_entry_ids.add(entry.entry_id)
        if any(item.method == "ISBN" for item in entry_results):
            exact_isbn_ids.add(entry.entry_id)
        if any(item.method in {"Название", "Название и автор"} for item in entry_results):
            exact_title_ids.add(entry.entry_id)
        if any(item.status == "Возможное совпадение" for item in entry_results):
            probable_ids.add(entry.entry_id)

    return results, matched_entry_ids, exact_isbn_ids, exact_title_ids, probable_ids


class DatabaseIndex:
    def __init__(self, records: list[DatabaseRecord]) -> None:
        self.records = records
        self.by_isbn: dict[str, list[DatabaseRecord]] = defaultdict(list)
        self.by_title: dict[str, list[DatabaseRecord]] = defaultdict(list)

        for record in records:
            for isbn in record.isbns:
                for normalized in extract_isbns(isbn):
                    self.by_isbn[normalized].append(record)
            for title in record.titles:
                normalized = normalize_title(title)
                if normalized:
                    self.by_title[normalized].append(record)

        self.unique_titles = list(self.by_title.keys())

    @staticmethod
    def _author_matches(entry_author: str, record: DatabaseRecord) -> bool:
        excel_identity = _author_identity(entry_author)
        if not excel_identity or not record.authors:
            return True
        excel_surname, excel_initials = excel_identity

        for author in record.authors:
            record_identity = _author_identity(author)
            if not record_identity:
                continue
            record_surname, record_initials = record_identity
            if record_surname != excel_surname:
                continue
            if excel_initials and record_initials:
                compare_count = min(len(excel_initials), len(record_initials))
                if excel_initials[:compare_count] != record_initials[:compare_count]:
                    continue
            return True
        return False

    @staticmethod
    def _author_similarity(entry_author: str, record: DatabaseRecord) -> float:
        if not entry_author or not record.authors:
            return 100.0
        entry_normalized = normalize_author(entry_author)
        if fuzz is None:
            return 100.0 if DatabaseIndex._author_matches(entry_author, record) else 0.0
        return max(fuzz.token_set_ratio(entry_normalized, normalize_author(author)) for author in record.authors)

    def match(
        self,
        entry: ExcelEntry,
        use_isbn_matching: bool,
        use_title_fallback: bool,
        use_fuzzy: bool,
        fuzzy_threshold: int,
    ) -> list[MatchResult]:
        normalized_isbns = extract_isbns(entry.isbn) if use_isbn_matching else []
        isbn_records: list[DatabaseRecord] = []
        seen_record_numbers: set[int] = set()
        if use_isbn_matching:
            for normalized_isbn in normalized_isbns:
                for record in self.by_isbn.get(normalized_isbn, []):
                    if record.record_number not in seen_record_numbers:
                        seen_record_numbers.add(record.record_number)
                        isbn_records.append(record)
        if isbn_records:
            return [
                MatchResult(
                    status="Совпадение",
                    method="ISBN",
                    confidence=100.0,
                    excel=entry,
                    database=record,
                    source_type=SOURCE_SUBSTANCES,
                    matched_value=entry.isbn,
                )
                for record in isbn_records
            ]

        if not use_title_fallback:
            if not use_isbn_matching:
                note = "Поиск по ISBN и названию отключён"
            else:
                note = "Корректный ISBN не найден" if entry.isbn and not normalized_isbns else "Совпадение по ISBN отсутствует"
            return [MatchResult("Не найдено", "—", 0.0, entry, note=note)]

        normalized_title = normalize_title(entry.title)
        if not normalized_title:
            if entry.isbn and not normalized_isbns:
                note = "ISBN имеет неверный формат или контрольную цифру; названия для резервного поиска нет"
            else:
                note = "Нет ISBN или названия для поиска"
            return [MatchResult("Не найдено", "—", 0.0, entry, note=note)]

        exact_candidates = self.by_title.get(normalized_title, [])
        if exact_candidates:
            author_candidates = [
                record for record in exact_candidates if self._author_matches(entry.author, record)
            ]
            if author_candidates:
                return [
                    MatchResult(
                        status="Совпадение",
                        method="Название и автор" if entry.author else "Название",
                        confidence=100.0,
                        excel=entry,
                        database=record,
                        source_type=SOURCE_SUBSTANCES,
                        matched_value=entry.title or entry.author,
                    )
                    for record in author_candidates
                ]

        if use_fuzzy and process is not None and self.unique_titles:
            candidate_titles = process.extract(
                normalized_title,
                self.unique_titles,
                scorer=fuzz.token_set_ratio,
                score_cutoff=fuzzy_threshold,
                limit=5,
            )
            best: list[tuple[float, DatabaseRecord, float, float]] = []
            for title_key, title_score, _ in candidate_titles:
                for record in self.by_title[title_key]:
                    author_score = self._author_similarity(entry.author, record)
                    if entry.author and record.authors and author_score < 60:
                        continue
                    combined = title_score * 0.85 + author_score * 0.15
                    if combined >= fuzzy_threshold:
                        best.append((combined, record, float(title_score), float(author_score)))

            if best:
                best.sort(key=lambda item: item[0], reverse=True)
                best_score = best[0][0]
                selected = [item for item in best if item[0] >= best_score - 1.0][:3]
                return [
                    MatchResult(
                        status="Возможное совпадение",
                        method="Приблизительно по названию и автору",
                        confidence=round(score, 1),
                        excel=entry,
                        database=record,
                        note=f"Название: {title_score:.0f}%, автор: {author_score:.0f}%",
                        source_type=SOURCE_SUBSTANCES,
                        matched_value=entry.title or entry.author,
                    )
                    for score, record, title_score, author_score in selected
                ]

        note = "ISBN не найден; название и автор не совпали" if normalized_isbns else "Название и автор не совпали"
        return [MatchResult("Не найдено", "—", 0.0, entry, note=note)]


def compare_database_records(
    records: list[DatabaseRecord],
    excel_paths: list[str | Path],
    *,
    database_label: str = "ИРБИС",
    foreign_agents_path: str | Path | None = None,
    use_isbn_matching: bool = True,
    use_title_fallback: bool = True,
    use_fuzzy: bool = False,
    fuzzy_threshold: int = 90,
    progress_cb: Optional[ProgressCallback] = None,
    cancel_cb: Optional[CancelCallback] = None,
) -> tuple[list[MatchResult], ComparisonSummary]:
    """Сверяет уже загруженные записи, в том числе полученные напрямую с ИРБИС."""
    _cancelled(cancel_cb)
    index = DatabaseIndex(records)
    entries, warnings = read_excel_entries(excel_paths, progress_cb, cancel_cb)
    foreign_entries, foreign_warnings = read_foreign_agent_entries(
        foreign_agents_path,
        progress_cb,
        cancel_cb,
    )
    warnings.extend(foreign_warnings)

    if entries and foreign_entries and progress_cb:
        progress_cb(53, "Параллельная сверка по веществам и иностранным агентам")

    if entries and foreign_entries:
        with ThreadPoolExecutor(max_workers=2) as executor:
            substance_future = executor.submit(
                compare_substance_entries,
                index,
                entries,
                use_isbn_matching,
                use_title_fallback,
                use_fuzzy,
                fuzzy_threshold,
                progress_cb,
                cancel_cb,
            )
            foreign_future = executor.submit(
                compare_foreign_agents,
                records,
                foreign_entries,
                progress_cb,
                cancel_cb,
            )
            substance_results, matched_entry_ids, exact_isbn_ids, exact_title_ids, probable_ids = substance_future.result()
            foreign_results = foreign_future.result()
    else:
        (
            substance_results,
            matched_entry_ids,
            exact_isbn_ids,
            exact_title_ids,
            probable_ids,
        ) = compare_substance_entries(
            index,
            entries,
            use_isbn_matching,
            use_title_fallback,
            use_fuzzy,
            fuzzy_threshold,
            progress_cb,
            cancel_cb,
        )
        foreign_results = compare_foreign_agents(records, foreign_entries, progress_cb, cancel_cb)

    results: list[MatchResult] = []
    results.extend(substance_results)
    results.extend(foreign_results)
    matched_foreign_ids = {
        result.foreign_agent.entry_id
        for result in foreign_results
        if result.foreign_agent is not None and result.status == "Совпадение"
    }
    substance_records = {
        result.database.record_number
        for result in results
        if result.status == "Совпадение"
        and result.database is not None
        and result.source_type == SOURCE_SUBSTANCES
    }
    foreign_records = {
        result.database.record_number
        for result in foreign_results
        if result.status == "Совпадение" and result.database is not None
    }

    summary = ComparisonSummary(
        database_file=database_label,
        excel_files=[str(Path(path)) for path in excel_paths],
        database_records=len(records),
        database_records_with_isbn=sum(1 for record in records if any(extract_isbns(value) for value in record.isbns)),
        excel_rows=len(entries),
        matched_excel_rows=len(matched_entry_ids),
        unmatched_excel_rows=len(entries) - len(matched_entry_ids),
        result_rows=len(results),
        exact_isbn_rows=len(exact_isbn_ids),
        exact_title_rows=len(exact_title_ids),
        probable_rows=len(probable_ids),
        foreign_agents_file=str(Path(foreign_agents_path)) if foreign_agents_path else "",
        foreign_agent_rows=len(foreign_entries),
        matched_foreign_agent_rows=len(matched_foreign_ids),
        foreign_agent_result_rows=len(foreign_results),
        substance_matched_records=len(substance_records),
        foreign_agent_matched_records=len(foreign_records),
        warnings=warnings,
    )

    if progress_cb:
        progress_cb(82, "Сравнение завершено, подготовка отчёта")
    return results, summary


def compare_files(
    database_path: str | Path | list[str | Path],
    excel_paths: list[str | Path],
    *,
    foreign_agents_path: str | Path | None = None,
    use_isbn_matching: bool = True,
    use_title_fallback: bool = True,
    use_fuzzy: bool = False,
    fuzzy_threshold: int = 90,
    progress_cb: Optional[ProgressCallback] = None,
    cancel_cb: Optional[CancelCallback] = None,
) -> tuple[list[MatchResult], ComparisonSummary]:
    database_paths = database_path if isinstance(database_path, list) else [database_path]
    records: list[DatabaseRecord] = []
    for source in database_paths:
        source_records = parse_database(source, progress_cb, cancel_cb)
        for record in source_records:
            record.record_number = len(records) + 1
            records.append(record)
    return compare_database_records(
        records,
        excel_paths,
        database_label="; ".join(str(Path(path)) for path in database_paths),
        foreign_agents_path=foreign_agents_path,
        use_isbn_matching=use_isbn_matching,
        use_title_fallback=use_title_fallback,
        use_fuzzy=use_fuzzy,
        fuzzy_threshold=fuzzy_threshold,
        progress_cb=progress_cb,
        cancel_cb=cancel_cb,
    )


def _publication_text(record: Optional[DatabaseRecord]) -> str:
    if not record:
        return ""
    parts: list[str] = []
    for item in record.publication:
        city = _extract_subfield(item, "A")
        publisher = _extract_subfield(item, "C")
        year = _extract_subfield(item, "D")
        publication = ", ".join(value for value in (city, publisher, year) if value)
        if publication:
            parts.append(publication)
    return " | ".join(dict.fromkeys(parts))


def _unique_join(values: Iterable[Any], separator: str = " | ") -> str:
    cleaned = [safe_text(value) for value in values if safe_text(value)]
    return separator.join(dict.fromkeys(cleaned))


def _matched_database_records(
    results: list[MatchResult],
    source_type: str | None = None,
) -> list[DatabaseRecord]:
    """Возвращает уникальные подтверждённые записи TXT для выбранного источника."""
    unique: dict[int, DatabaseRecord] = {}
    for result in results:
        if result.status != "Совпадение" or result.database is None:
            continue
        if source_type is not None and result.source_type != source_type:
            continue
        unique.setdefault(result.database.record_number, result.database)
    return [unique[key] for key in sorted(unique)]


COMMON_MATCH_HEADERS = [
    "№",
    "Автор",
    "Название",
    "ISBN",
    "Инвентарные номера",
    "Издание: город, издательство, год",
    "Номер записи в текстовой базе",
]

SUBSTANCE_MATCH_HEADERS = [
    *COMMON_MATCH_HEADERS,
    "Почему добавлено",
    "Файл-источник",
]

FOREIGN_AGENT_MATCH_HEADERS = [
    *COMMON_MATCH_HEADERS,
    "Совпавшее поле TXT",
    "Совпавшее значение",
    "№ в реестре",
    "Иностранный агент",
    "Тип иностранного агента",
    "Дата включения",
    "Вид записи в реестре",
    "Файл реестра",
    "Лист",
    "Строка",
]

COMBINED_MATCH_HEADERS = [
    *COMMON_MATCH_HEADERS,
    "Список совпадений",
    "Способ совпадения",
    "Совпавшее значение",
]


def _confirmed_results_by_record(
    results: list[MatchResult],
    source_type: str | None = None,
) -> dict[int, list[MatchResult]]:
    grouped: dict[int, list[MatchResult]] = defaultdict(list)
    for result in results:
        if result.status != "Совпадение" or result.database is None:
            continue
        if source_type is not None and result.source_type != source_type:
            continue
        grouped[result.database.record_number].append(result)
    return grouped


def _report_records(
    results: list[MatchResult],
    source_type: str | None = None,
    *,
    deduplicate: bool = False,
    sort_by: str = "record",
) -> tuple[list[DatabaseRecord], dict[int, list[MatchResult]]]:
    """Prepares report rows without changing matching or marker behavior."""
    records = _matched_database_records(results, source_type)
    grouped = _confirmed_results_by_record(results, source_type)

    if deduplicate:
        unique: dict[tuple[Any, ...], DatabaseRecord] = {}
        merged: dict[int, list[MatchResult]] = defaultdict(list)
        for record in records:
            normalized_isbns = [isbn for value in record.isbns for isbn in extract_isbns(value)]
            if normalized_isbns:
                identity: tuple[Any, ...] = ("isbn", normalized_isbns[0])
            else:
                title = normalize_title(next(iter(record.titles), ""))
                author = normalize_author(next(iter(record.authors), ""))
                identity = ("title_author", title, author) if title or author else ("record", record.record_number)
            selected = unique.setdefault(identity, record)
            merged[selected.record_number].extend(grouped.get(record.record_number, []))
        records = list(unique.values())
        grouped = dict(merged)

    def sort_key(record: DatabaseRecord) -> tuple[Any, ...]:
        if sort_by == "title":
            value = normalize_title(next(iter(record.titles), ""))
        elif sort_by == "author":
            value = normalize_author(next(iter(record.authors), ""))
        elif sort_by == "isbn":
            value = next((isbn for raw in record.isbns for isbn in extract_isbns(raw)), "")
        else:
            return (record.record_number,)
        return (not bool(value), value, record.record_number)

    records.sort(key=sort_key)
    return records, grouped


def _parallel_join(values: Iterable[Any]) -> str:
    """Объединяет значения построчно, сохраняя соответствие между колонками."""
    return "\n".join(safe_text(value) for value in values)


def _base_txt_row(number: int, record: DatabaseRecord) -> list[Any]:
    return [
        number,
        _unique_join(record.authors),
        _unique_join(record.titles),
        _unique_join(record.isbns),
        _unique_join(record.inventory_numbers),
        _publication_text(record),
        record.record_number,
    ]


def _substance_match_row(
    number: int,
    record: DatabaseRecord,
    record_results: list[MatchResult],
) -> list[Any]:
    reasons = []
    for result in record_results:
        method = safe_text(result.method).strip()
        matched_value = safe_text(result.matched_value).strip()
        reasons.append(
            f"{method} — {matched_value}" if method and matched_value else method or matched_value
        )
    return [
        *_base_txt_row(number, record),
        _unique_join(reasons, "\n"),
        _unique_join(
            (Path(result.excel.source_file).name for result in record_results),
            "\n",
        ),
    ]


def _foreign_agent_field(result: MatchResult) -> str:
    if ":" in result.method:
        return result.method.rsplit(":", 1)[-1].strip()
    return result.method


def _foreign_agent_match_row(
    number: int,
    record: DatabaseRecord,
    record_results: list[MatchResult],
) -> list[Any]:
    entries = [result.foreign_agent for result in record_results]
    return [
        *_base_txt_row(number, record),
        _parallel_join(_foreign_agent_field(result) for result in record_results),
        _parallel_join(result.matched_value for result in record_results),
        _parallel_join(entry.registry_number if entry else "" for entry in entries),
        _parallel_join(entry.name if entry else "" for entry in entries),
        _parallel_join(entry.agent_type if entry else "" for entry in entries),
        _parallel_join(entry.inclusion_date if entry else "" for entry in entries),
        _parallel_join(result.note for result in record_results),
        _parallel_join(Path(entry.source_file).name if entry else "" for entry in entries),
        _parallel_join(entry.sheet_name if entry else "" for entry in entries),
        _parallel_join(entry.row_number if entry else "" for entry in entries),
    ]


def _combined_match_row(
    number: int,
    record: DatabaseRecord,
    record_results: list[MatchResult],
) -> list[Any]:
    return [
        *_base_txt_row(number, record),
        _parallel_join(result.source_type for result in record_results),
        _parallel_join(result.method for result in record_results),
        _parallel_join(result.matched_value for result in record_results),
    ]


def _style_match_sheet(
    worksheet,
    row_count: int,
    headers: list[str],
    widths: dict[str, int],
    row_fill_color: str,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    row_fill = PatternFill("solid", fgColor=row_fill_color)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{max(row_count, 1)}"
    )
    worksheet.row_dimensions[1].height = 42

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, max_row=row_count), start=2
    ):
        max_lines = max(
            (safe_text(cell.value).count("\n") + 1 for cell in row),
            default=1,
        )
        worksheet.row_dimensions[row_number].height = min(120, max(30, 18 * max_lines))
        for cell in row:
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _add_results_sheet(
    workbook,
    *,
    title: str,
    headers: list[str],
    records: list[DatabaseRecord],
    grouped_results: dict[int, list[MatchResult]],
    row_builder,
    table_name: str,
    table_style: str,
    row_fill_color: str,
    widths: dict[str, int],
    active: bool = False,
    cancel_cb: Optional[CancelCallback] = None,
):
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    worksheet = workbook.active if active else workbook.create_sheet()
    worksheet.title = title
    worksheet.append(headers)

    for number, record in enumerate(records, start=1):
        if number % 500 == 0:
            _cancelled(cancel_cb)
        worksheet.append(
            row_builder(number, record, grouped_results.get(record.record_number, []))
        )

    _style_match_sheet(
        worksheet,
        len(records) + 1,
        headers,
        widths,
        row_fill_color,
    )

    if records:
        last_column = get_column_letter(len(headers))
        table = Table(
            displayName=table_name,
            ref=f"A1:{last_column}{len(records) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    return worksheet


def _add_summary_sheet(workbook, summary: ComparisonSummary, *, active: bool = False):
    from openpyxl.styles import Font, PatternFill

    worksheet = workbook.active if active else workbook.create_sheet()
    worksheet.title = "Сводка"
    worksheet.append(["Показатель", "Значение"])
    rows = [
        ("Источник библиографических записей", summary.database_file),
        ("Проверено библиографических записей", summary.database_records),
        ("Строк в перечнях веществ", summary.excel_rows),
        ("Найдено строк перечней веществ", summary.matched_excel_rows),
        ("Записей с совпадениями по веществам", summary.substance_matched_records),
        ("Строк в реестре иностранных агентов", summary.foreign_agent_rows),
        ("Найдено строк реестра иностранных агентов", summary.matched_foreign_agent_rows),
        ("Записей с совпадениями по иностранным агентам", summary.foreign_agent_matched_records),
    ]
    for row in rows:
        worksheet.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    worksheet.column_dimensions["A"].width = 48
    worksheet.column_dimensions["B"].width = 64
    worksheet.freeze_panes = "A2"
    return worksheet


def export_results(
    output_path: str | Path,
    results: list[MatchResult],
    summary: ComparisonSummary,
    progress_cb: Optional[ProgressCallback] = None,
    cancel_cb: Optional[CancelCallback] = None,
    report_options: dict[str, Any] | None = None,
) -> Path:
    """Создаёт выбранные пользователем листы с точными совпадениями."""
    from openpyxl import Workbook

    options = {
        "substances": True,
        "foreign_agents": True,
        "combined": False,
        "summary": False,
        "deduplicate": False,
        "sort": "record",
    }
    if report_options:
        for key in ("substances", "foreign_agents", "combined", "summary", "deduplicate"):
            if key in report_options:
                options[key] = bool(report_options[key])
        if report_options.get("sort") in {"record", "title", "author", "isbn"}:
            options["sort"] = str(report_options["sort"])
    if not any(bool(options[key]) for key in ("substances", "foreign_agents", "combined", "summary")):
        raise ValueError("Для Excel-отчёта должен быть выбран хотя бы один лист.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if progress_cb:
        progress_cb(85, "Создание выбранных списков точных совпадений")

    report_kwargs = {
        "deduplicate": bool(options["deduplicate"]),
        "sort_by": str(options["sort"]),
    }
    substance_records, substance_grouped = _report_records(results, SOURCE_SUBSTANCES, **report_kwargs)
    foreign_agent_records, foreign_agent_grouped = _report_records(results, SOURCE_FOREIGN_AGENTS, **report_kwargs)
    combined_records, combined_grouped = _report_records(results, **report_kwargs)

    workbook = Workbook()
    sheet_created = False
    if options["substances"]:
        _add_results_sheet(
            workbook,
            title="Вещества",
            headers=SUBSTANCE_MATCH_HEADERS,
            records=substance_records,
            grouped_results=substance_grouped,
            row_builder=_substance_match_row,
            table_name="SubstanceMatchesTable",
            table_style="TableStyleMedium4",
            row_fill_color="E2F0D9",
            widths={
                "A": 7, "B": 34, "C": 52, "D": 24, "E": 24, "F": 42,
                "G": 24, "H": 48, "I": 34,
            },
            active=not sheet_created,
            cancel_cb=cancel_cb,
        )
        sheet_created = True
    if options["foreign_agents"]:
        _add_results_sheet(
            workbook,
            title="Иностранные агенты",
            headers=FOREIGN_AGENT_MATCH_HEADERS,
            records=foreign_agent_records,
            grouped_results=foreign_agent_grouped,
            row_builder=_foreign_agent_match_row,
            table_name="ForeignAgentMatchesTable",
            table_style="TableStyleMedium9",
            row_fill_color="DDEBF7",
            widths={
                "A": 7, "B": 34, "C": 52, "D": 24, "E": 24, "F": 42,
                "G": 24, "H": 24, "I": 38, "J": 16, "K": 48, "L": 28,
                "M": 18, "N": 24, "O": 34, "P": 22, "Q": 12,
            },
            active=not sheet_created,
            cancel_cb=cancel_cb,
        )
        sheet_created = True
    if options["combined"]:
        _add_results_sheet(
            workbook,
            title="Все совпадения",
            headers=COMBINED_MATCH_HEADERS,
            records=combined_records,
            grouped_results=combined_grouped,
            row_builder=_combined_match_row,
            table_name="AllMatchesTable",
            table_style="TableStyleMedium2",
            row_fill_color="FFF2CC",
            widths={
                "A": 7, "B": 34, "C": 52, "D": 24, "E": 24,
                "F": 42, "G": 24, "H": 28, "I": 34, "J": 42,
            },
            active=not sheet_created,
            cancel_cb=cancel_cb,
        )
        sheet_created = True
    if options["summary"]:
        _add_summary_sheet(workbook, summary, active=not sheet_created)
        sheet_created = True

    workbook.active = 0
    workbook.save(output_path)
    summary.output_file = str(output_path)
    if progress_cb:
        progress_cb(
            92,
            f"Excel-отчёт создан: {output_path.name}. "
            f"Вещества: {len(substance_records):,}; "
            f"иноагенты: {len(foreign_agent_records):,}",
        )
    return output_path

def _detect_newline(text: str) -> str:
    if "\r\n" in text:
        return "\r\n"
    if "\n" in text:
        return "\n"
    if "\r" in text:
        return "\r"
    return "\n"


def _field_number(line: str) -> Optional[int]:
    """Возвращает номер поля строки вида ``#333: ...``."""
    match = re.match(r"^\s*#(\d+):", line)
    return int(match.group(1)) if match else None


def _ordered_field_insert_index(lines: list[str], target_field: int) -> int:
    """Находит вертикальную позицию поля в числовом порядке записи.

    Новое поле ставится после всех полей с номером не больше целевого и
    перед первым полем с большим номером. Благодаря этому #333 не уезжает
    к #900, а отсутствующее #900 появляется перед #910 и последующими полями.
    """
    last_field_index: Optional[int] = None
    for index, line in enumerate(lines):
        field_number = _field_number(line)
        if field_number is None:
            continue
        if field_number > target_field:
            return index
        last_field_index = index

    # Не помещаем новое поле после служебных пустых строк в конце записи.
    return last_field_index + 1 if last_field_index is not None else len(lines)


def _record_author_marker(raw_record: str) -> str:
    authors: list[str] = []
    for line in re.split(r"\r\n|\n|\r", raw_record):
        match = re.match(r"^\s*#(?:700|701|702):\s*(.*)$", line)
        if not match:
            continue
        value = match.group(1)
        parts = [
            _extract_subfield(value, "A"),
            _extract_subfield(value, "B"),
            _extract_subfield(value, "G"),
        ]
        author = " ".join(part for part in parts if part).strip()
        if author:
            authors.append(author)
    if not authors:
        return ""
    return _unique_join(authors, "; ").upper()


def _registry_person_marker_name(value: str) -> str:
    """Формирует человекочитаемое ФИО для метки иностранного агента.

    Реестр может хранить псевдоним либо как ``(псевдоним: ...)``, либо в
    кавычках после ФИО. В библиотечную запись записываем только персону, а
    псевдоним явно подписываем, чтобы он не выглядел второй фамилией/частью ФИО.
    """
    raw = unicodedata.normalize("NFKC", safe_text(value)).strip()
    if not raw:
        return ""

    explicit_aliases = re.findall(
        r"\(\s*псевдоним\s*:\s*([^)]+)\)", raw, flags=re.IGNORECASE
    )
    quoted_aliases = re.findall(r'[«"]([^»"]+)[»"]', raw)

    # Для персональных записей текст в кавычках в реестре используется как
    # псевдоним. Убираем его из основного ФИО и выводим отдельно.
    main_raw = re.sub(
        r"\(\s*псевдоним\s*:\s*[^)]+\)", " ", raw, flags=re.IGNORECASE
    )
    main_raw = re.sub(r'[«"]([^»"]+)[»"]', ' ', main_raw)
    main_name = _registry_plain_name(main_raw)
    if not main_name:
        main_name = _registry_plain_name(raw)

    aliases: list[str] = []
    for alias_source in [*explicit_aliases, *quoted_aliases]:
        for part in re.split(r"\s*[,;]\s*", alias_source):
            clean = _registry_plain_name(part)
            if clean and normalize_author(clean) != normalize_author(main_name):
                aliases.append(clean)
    aliases = list(dict.fromkeys(aliases))

    main_display = re.sub(r"\s+", " ", main_name).strip().upper()
    alias_display = [re.sub(r"\s+", " ", item).strip().upper() for item in aliases]
    if main_display and alias_display:
        return f"{main_display} (ПСЕВДОНИМ: {'; '.join(alias_display)})"
    return main_display


def _foreign_agent_marker_name(result: MatchResult) -> str:
    """Возвращает именно совпавшего автора, а не название родительской записи.

    Особенно важно для реестровых записей-организаций: если книга совпала с
    человеком из списка участников, метка должна содержать этого человека, а
    не, например, название проекта/организации «НАСТОЯЩАЯ РОССИЯ».
    """
    entry = result.foreign_agent
    matched = re.sub(r"\s+", " ", result.matched_value).strip() if result.matched_value else ""

    if entry is not None:
        # Совпадение по участнику: находим исходную строку участника, чтобы не
        # потерять псевдоним, и форматируем именно её.
        if result.note.startswith("Участник"):
            matched_normalized = normalize_author(matched)
            for participant in entry.participants:
                variants = _registry_name_variants(participant, is_person=True)
                if any(normalize_author(item) == matched_normalized for item in variants):
                    return _registry_person_marker_name(participant)
            if matched:
                return _registry_person_marker_name(matched)

        # Основная запись физического лица: используем полное ФИО из реестра,
        # включая правильно оформленный псевдоним.
        if "физичес" in normalize_author(entry.agent_type) and entry.name:
            return _registry_person_marker_name(entry.name)

    # Резервный вариант — именно совпавшее значение автора, но не название
    # родительской организации/проекта.
    if matched:
        return _registry_person_marker_name(matched)
    return _record_author_marker(result.database.raw_record) if result.database else ""


def _result_is_eligible_for_txt_marker(result: MatchResult) -> bool:
    """Для иноагентов разрешает TXT-пометку только при совпадении по автору."""
    if result.status != "Совпадение" or result.database is None:
        return False
    if result.source_type != SOURCE_FOREIGN_AGENTS:
        return True
    return result.method == "Реестр иностранных агентов: Автор"


def _normalized_marker_text(value: str) -> str:
    # ИРБИС/Excel иногда приносят визуально невидимые Unicode-символы
    # (WORD JOINER, zero-width, soft hyphen и т. п.). На экране две метки
    # выглядят одинаково, но простое сравнение строк считает их разными.
    # Для дедупликации убираем все форматирующие/управляющие символы и
    # приводим любые Unicode-разделители к обычному пробелу.
    # Некоторые источники сохраняют пробел как буквальную HTML-сущность
    # ``&#x20;``. Для сравнения меток это тот же обычный пробел.
    normalized = unicodedata.normalize("NFKC", html.unescape(safe_text(value)))
    cleaned: list[str] = []
    for char in normalized:
        category = unicodedata.category(char)
        if category.startswith("C"):
            continue
        if category.startswith("Z"):
            cleaned.append(" ")
        else:
            cleaned.append(char)
    normalized = "".join(cleaned).replace("ё", "е").replace("Ё", "Е")
    return re.sub(r"\s+", " ", normalized).strip().casefold()


def _foreign_marker_identity(value: str) -> str | None:
    """Ключ короткой авторской метки ^AI^@... для удаления повторов.

    Если в метке явно указан псевдоним, используем его как идентификатор
    персоны. В реестре встречаются дубли строк с опечаткой в настоящей
    фамилии, но с одним и тем же псевдонимом (например, ЧХАРТИШВИЛИ и
    ЧХАРТИШВИЛЛИ при псевдониме БОРИС АКУНИН). Для библиотечной записи это
    одна и та же отметка, а не два разных иностранных агента.

    В остальных случаях возвращаем нормализованное имя после ^AI^@. Это позволяет
    схлопывать два визуально одинаковых повторения даже при скрытых Unicode
    символах, попавших из реестра/Excel. Другие значения поля 333 не трогаем.
    """
    normalized = _normalized_marker_text(value)
    prefix = "^ai^@"
    if not normalized.startswith(prefix):
        return None
    name = normalized[len(prefix):].strip()
    if not name:
        return None
    pseudonym_match = re.search(r"\(\s*псевдоним\s*:\s*([^)]+)\)", name, flags=re.IGNORECASE)
    if pseudonym_match:
        pseudonym = re.sub(
            r"[^0-9a-zа-я]+",
            "",
            pseudonym_match.group(1),
            flags=re.IGNORECASE,
        )
        if pseudonym:
            return pseudonym
    # Пунктуация и пробелы в ФИО/псевдониме не должны превращать одного
    # автора в две отдельные метки. Буквы и цифры сохраняем.
    return re.sub(r"[^0-9a-zа-я]+", "", name, flags=re.IGNORECASE)


def _marker_only_repeat_count(value: str, marker: str) -> int:
    """Возвращает число повторов, если поле состоит только из одной метки N раз."""
    normalized_value = _normalized_marker_text(value)
    normalized_marker = _normalized_marker_text(marker)
    if not normalized_marker or not normalized_value:
        return 0
    count = 0
    remaining = normalized_value
    while remaining.startswith(normalized_marker):
        count += 1
        remaining = remaining[len(normalized_marker):].strip()
    return count if count and not remaining else 0


def _field_contains_marker(value: str, marker: str) -> bool:
    """Проверяет наличие метки в поле, включая поле с дополнительными подполями."""
    normalized_value = _normalized_marker_text(value)
    normalized_marker = _normalized_marker_text(marker)
    if not normalized_marker:
        return True
    if normalized_marker.startswith("^"):
        return bool(re.search(re.escape(normalized_marker) + r"(?=\^|$)", normalized_value))
    return bool(
        re.search(
            r"(?<!\w)" + re.escape(normalized_marker) + r"(?!\w)",
            normalized_value,
        )
    )


def _modify_matched_record(
    raw_record: str,
    newline: str,
    markers_333: Iterable[str],
    age_marker: str = DEFAULT_AGE_MARKER,
    *,
    marker_field: int = DEFAULT_SUBSTANCE_MARKER_FIELD,
    age_field: int = DEFAULT_AGE_MARKER_FIELD,
    field_markers: Iterable[tuple[int, str]] | None = None,
) -> tuple[str, bool]:
    """Добавляет метки в правильных вертикальных позициях без дублирования."""
    had_trailing_newline = raw_record.endswith(("\r\n", "\n", "\r"))
    lines = re.split(r"\r\n|\n|\r", raw_record)
    if had_trailing_newline and lines and lines[-1] == "":
        lines = lines[:-1]

    changed = False

    requested_markers = list(
        field_markers
        if field_markers is not None
        else ((marker_field, marker) for marker in markers_333)
    )
    markers: list[tuple[int, str]] = []
    marker_keys: set[tuple[int, str]] = set()
    for field_number, marker in requested_markers:
        marker = marker.strip() if marker else ""
        if not marker:
            continue
        identity = _foreign_marker_identity(marker)
        key = (
            int(field_number),
            f"foreign:{identity}" if identity is not None else _normalized_marker_text(marker),
        )
        if key in marker_keys:
            continue
        marker_keys.add(key)
        markers.append((int(field_number), marker))
    requested_foreign_keys = {
        (field_number, identity)
        for field_number, marker in markers
        if (identity := _foreign_marker_identity(marker)) is not None
    }
    if requested_foreign_keys:
        seen_foreign: set[tuple[int, str]] = set()
        deduped_lines: list[str] = []
        for line in lines:
            field_number = _field_number(line)
            match = re.match(r"^\s*#\d{1,3}:\s*(.*?)\s*$", line)
            identity = _foreign_marker_identity(match.group(1)) if match else None
            pair = (field_number, identity) if field_number is not None and identity is not None else None
            if pair is not None and pair in requested_foreign_keys:
                if pair in seen_foreign:
                    changed = True
                    continue
                seen_foreign.add(pair)
            deduped_lines.append(line)
        lines = deduped_lines

    # Если в TXT уже лежит несколько одинаковых повторений нашей метки,
    # исправляем их так же, как в прямом режиме ИРБИС.
    for field_number, marker in markers:
        matching_indices: list[int] = []
        for index, line in enumerate(lines):
            if _field_number(line) != field_number:
                continue
            match = re.match(r"^\s*#\d{1,3}:\s*(.*?)\s*$", line)
            if not match or not _field_contains_marker(match.group(1), marker):
                continue
            value = match.group(1)
            if _marker_only_repeat_count(value, marker) > 1:
                lines[index] = f"#{field_number:03d}: {marker}"
                changed = True
            matching_indices.append(index)

        if len(matching_indices) > 1:
            for index in reversed(matching_indices[1:]):
                match = re.match(r"^\s*#\d{1,3}:\s*(.*?)\s*$", lines[index])
                if match and _marker_only_repeat_count(match.group(1), marker) == 1:
                    del lines[index]
                    changed = True

    existing_field_values: dict[int, list[str]] = defaultdict(list)
    for line in lines:
        field_number = _field_number(line)
        match = re.match(r"^\s*#\d{1,3}:\s*(.*?)\s*$", line)
        if field_number is not None and match:
            existing_field_values[field_number].append(match.group(1))

    # Возрастная метка дописывается в каждое существующее выбранное поле.
    has_age_field = False
    age_pattern = re.compile(rf"^(\s*#{age_field:03d}:\s*)(.*?)(\s*)$")
    for index, line in enumerate(lines):
        match = age_pattern.match(line)
        if not match:
            continue
        has_age_field = True
        prefix, value, trailing = match.groups()
        if age_marker and not _field_contains_marker(value, age_marker):
            lines[index] = f"{prefix}{value}{age_marker}{trailing}"
            changed = True

    for field_number, marker in markers:
        if not any(
            _field_contains_marker(value, marker)
            for value in existing_field_values[field_number]
        ):
            lines.insert(
                _ordered_field_insert_index(lines, field_number),
                f"#{field_number:03d}: {marker}",
            )
            existing_field_values[field_number].append(marker)
            changed = True

    if age_marker and not has_age_field:
        lines.insert(
            _ordered_field_insert_index(lines, age_field),
            f"#{age_field:03d}: {age_marker}",
        )
        changed = True

    modified = newline.join(lines)
    if had_trailing_newline:
        modified += newline
    return modified, changed


def _marker_removal_pattern(marker: str, *, name_template: bool = False) -> re.Pattern[str] | None:
    marker = marker.strip()
    if not marker:
        return None
    if name_template and "{name}" in marker:
        parts = marker.split("{name}")
        expression = r"[^\^\r\n]*".join(re.escape(part) for part in parts)
    else:
        expression = re.escape(marker)
    if marker.startswith("^"):
        expression += r"(?=\^|$)"
    else:
        expression = r"(?<!\w)" + expression + r"(?!\w)"
    return re.compile(expression, re.IGNORECASE)


def _record_contains_removal_marker(
    raw_record: str,
    removals: dict[int, list[tuple[str, bool]]],
) -> bool:
    for line in re.split(r"\r\n|\n|\r", raw_record):
        field_number = _field_number(line)
        patterns = removals.get(field_number or -1, [])
        match = re.match(r"^\s*#\d{1,3}:\s*(.*?)\s*$", line)
        if not patterns or not match:
            continue
        value = match.group(1)
        for marker, is_template in patterns:
            pattern = _marker_removal_pattern(marker, name_template=is_template)
            if pattern is not None and pattern.search(value):
                return True
    return False


def _remove_markers_from_record(
    raw_record: str,
    newline: str,
    removals: dict[int, list[tuple[str, bool]]],
    *,
    conditional_removals: dict[int, list[tuple[str, bool]]] | None = None,
    conditional_on: dict[int, list[tuple[str, bool]]] | None = None,
) -> tuple[str, bool]:
    """Удаляет только известные пометки, сохраняя прочее содержимое полей."""
    had_trailing_newline = raw_record.endswith(("\r\n", "\n", "\r"))
    lines = re.split(r"\r\n|\n|\r", raw_record)
    if had_trailing_newline and lines and lines[-1] == "":
        lines = lines[:-1]

    effective_removals: dict[int, list[tuple[str, bool]]] = {
        field_number: list(patterns)
        for field_number, patterns in removals.items()
    }
    if (
        conditional_removals
        and conditional_on
        and _record_contains_removal_marker(raw_record, conditional_on)
    ):
        for field_number, patterns in conditional_removals.items():
            effective_removals.setdefault(field_number, []).extend(patterns)

    changed = False
    cleaned_lines: list[str] = []
    for line in lines:
        field_number = _field_number(line)
        patterns = effective_removals.get(field_number or -1, [])
        match = re.match(r"^(\s*#\d{1,3}:\s*)(.*?)(\s*)$", line)
        if not patterns or not match:
            cleaned_lines.append(line)
            continue
        prefix, value, trailing = match.groups()
        cleaned_value = value
        for marker, is_template in patterns:
            pattern = _marker_removal_pattern(marker, name_template=is_template)
            if pattern is not None:
                cleaned_value = pattern.sub("", cleaned_value)
        if cleaned_value == value:
            cleaned_lines.append(line)
            continue
        changed = True
        if cleaned_value.strip():
            cleaned_lines.append(f"{prefix}{cleaned_value}{trailing}")

    modified = newline.join(cleaned_lines)
    if had_trailing_newline:
        modified += newline
    return modified, changed


def remove_markers_from_tag_values(
    tag_values: Iterable[tuple[int, str]],
    *,
    substance_marker: str = DEFAULT_SUBSTANCE_MARKER,
    foreign_agent_marker_template: str = DEFAULT_FOREIGN_AGENT_MARKER_TEMPLATE,
    age_marker: str = DEFAULT_AGE_MARKER,
    substance_marker_field: int = DEFAULT_SUBSTANCE_MARKER_FIELD,
    foreign_agent_marker_field: int = DEFAULT_FOREIGN_AGENT_MARKER_FIELD,
    age_marker_field: int = DEFAULT_AGE_MARKER_FIELD,
) -> tuple[list[tuple[int, str]], bool]:
    """Удаляет настроенные метки прямо из набора полей записи ИРБИС.

    В отличие от ``remove_database_markers`` эта функция не создаёт TXT-файл,
    поэтому подходит для прямой серверной очистки. Остальное содержимое полей
    и порядок повторений сохраняются.
    """
    values = [(int(tag), str(value)) for tag, value in tag_values]

    removals: dict[int, list[tuple[str, bool]]] = defaultdict(list)
    classification_removals: dict[int, list[tuple[str, bool]]] = defaultdict(list)
    substance_fields = {DEFAULT_SUBSTANCE_MARKER_FIELD, int(substance_marker_field)}
    foreign_fields = {DEFAULT_FOREIGN_AGENT_MARKER_FIELD, int(foreign_agent_marker_field)}
    age_fields = {DEFAULT_AGE_MARKER_FIELD, int(age_marker_field)}

    for field_number in substance_fields:
        patterns = [(DEFAULT_SUBSTANCE_MARKER, False), (substance_marker, False)]
        removals[field_number].extend(patterns)
        classification_removals[field_number].extend(patterns)
    for field_number in foreign_fields:
        patterns = [
            (DEFAULT_FOREIGN_AGENT_MARKER_TEMPLATE, True),
            (foreign_agent_marker_template, True),
        ]
        removals[field_number].extend(patterns)
        classification_removals[field_number].extend(patterns)
    for field_number in age_fields:
        removals[field_number].extend(
            [(DEFAULT_AGE_MARKER, False), (age_marker, False)]
        )

    # Совместимость со старой схемой: ^A18+ удаляется только из записи,
    # где присутствует наша классификационная метка.
    has_classification_marker = False
    for tag, value in values:
        for marker, is_template in classification_removals.get(tag, []):
            pattern = _marker_removal_pattern(marker, name_template=is_template)
            if pattern is not None and pattern.search(value):
                has_classification_marker = True
                break
        if has_classification_marker:
            break

    conditional_age_removals: dict[int, list[tuple[str, bool]]] = defaultdict(list)
    if has_classification_marker:
        for field_number in substance_fields | foreign_fields:
            conditional_age_removals[field_number].append(("^A18+", False))

    changed = False
    cleaned: list[tuple[int, str]] = []
    for tag, value in values:
        cleaned_value = value
        patterns = [*removals.get(tag, []), *conditional_age_removals.get(tag, [])]
        for marker, is_template in patterns:
            pattern = _marker_removal_pattern(marker, name_template=is_template)
            if pattern is not None:
                cleaned_value = pattern.sub("", cleaned_value)
        if cleaned_value != value:
            changed = True
        if cleaned_value.strip():
            cleaned.append((tag, cleaned_value))
        elif cleaned_value == value:
            # Пустое исходное поле не относится к очистке и должно сохраниться.
            cleaned.append((tag, value))

    return cleaned, changed


def remove_database_markers(
    source_path: str | Path,
    output_path: str | Path,
    *,
    substance_marker: str = DEFAULT_SUBSTANCE_MARKER,
    foreign_agent_marker_template: str = DEFAULT_FOREIGN_AGENT_MARKER_TEMPLATE,
    age_marker: str = DEFAULT_AGE_MARKER,
    substance_marker_field: int = DEFAULT_SUBSTANCE_MARKER_FIELD,
    foreign_agent_marker_field: int = DEFAULT_FOREIGN_AGENT_MARKER_FIELD,
    age_marker_field: int = DEFAULT_AGE_MARKER_FIELD,
) -> tuple[Path, int]:
    """Создаёт TXT-копию без стандартных и текущих настроенных пометок."""
    source = Path(source_path)
    output = Path(output_path)
    if source.resolve() == output.resolve():
        raise ValueError("Очищенная TXT-база не должна перезаписывать исходный файл.")

    removals: dict[int, list[tuple[str, bool]]] = defaultdict(list)
    classification_removals: dict[int, list[tuple[str, bool]]] = defaultdict(list)
    substance_fields = {DEFAULT_SUBSTANCE_MARKER_FIELD, substance_marker_field}
    foreign_fields = {DEFAULT_FOREIGN_AGENT_MARKER_FIELD, foreign_agent_marker_field}
    age_fields = {DEFAULT_AGE_MARKER_FIELD, age_marker_field}
    for field_number in substance_fields:
        patterns = [
            (DEFAULT_SUBSTANCE_MARKER, False),
            (substance_marker, False),
        ]
        removals[field_number].extend(patterns)
        classification_removals[field_number].extend(patterns)
    for field_number in foreign_fields:
        patterns = [
            (DEFAULT_FOREIGN_AGENT_MARKER_TEMPLATE, True),
            (foreign_agent_marker_template, True),
        ]
        removals[field_number].extend(patterns)
        classification_removals[field_number].extend(patterns)
    for field_number in age_fields:
        removals[field_number].extend([
            (DEFAULT_AGE_MARKER, False),
            (age_marker, False),
        ])

    # Старая пометка ^A18+ удаляется только из записей, где действительно была
    # пометка вещества или иноагента. Самостоятельные ^A18+ в базе сохраняются.
    conditional_age_removals: dict[int, list[tuple[str, bool]]] = defaultdict(list)
    for field_number in substance_fields | foreign_fields:
        conditional_age_removals[field_number].append(("^A18+", False))

    text, encoding = _read_text_file_with_encoding(source)
    newline = _detect_newline(text)
    parts = re.split(r"(\r?\n\*{5}\s*(?:\r?\n|$))", text)
    cleaned_records = 0
    for part_index in range(0, len(parts), 2):
        raw_record = parts[part_index]
        if not raw_record.strip():
            continue
        cleaned, changed = _remove_markers_from_record(
            raw_record,
            newline,
            removals,
            conditional_removals=conditional_age_removals,
            conditional_on=classification_removals,
        )
        parts[part_index] = cleaned
        if changed:
            cleaned_records += 1

    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding=encoding, newline="") as output_file:
        output_file.write("".join(parts))
    return output, cleaned_records


def build_markers_by_record(
    results: list[MatchResult],
    *,
    substance_marker: str = DEFAULT_SUBSTANCE_MARKER,
    foreign_agent_marker_template: str = DEFAULT_FOREIGN_AGENT_MARKER_TEMPLATE,
    substance_marker_field: int = DEFAULT_SUBSTANCE_MARKER_FIELD,
    foreign_agent_marker_field: int = DEFAULT_FOREIGN_AGENT_MARKER_FIELD,
) -> dict[int, list[tuple[int, str]]]:
    """Готовит метки по MFN/номеру записи для TXT и прямой записи в ИРБИС."""
    markers_by_record: dict[int, list[tuple[int, str]]] = defaultdict(list)
    for result in results:
        if not _result_is_eligible_for_txt_marker(result) or result.database is None:
            continue
        record_number = result.database.source_record_number or result.database.record_number
        if result.source_type == SOURCE_FOREIGN_AGENTS:
            agent_name = _foreign_agent_marker_name(result)
            marker = foreign_agent_marker_template.replace("{name}", agent_name)
            marker_field = foreign_agent_marker_field
        else:
            marker = substance_marker
            marker_field = substance_marker_field
        marker = marker.strip()
        field_marker = (int(marker_field), marker)
        foreign_identity = _foreign_marker_identity(marker) if result.source_type == SOURCE_FOREIGN_AGENTS else None
        marker_key = (
            int(marker_field),
            f"foreign:{foreign_identity}" if foreign_identity is not None else _normalized_marker_text(marker),
        )
        existing_keys = {
            (
                int(existing_field),
                f"foreign:{existing_identity}" if (existing_identity := _foreign_marker_identity(existing_marker)) is not None else _normalized_marker_text(existing_marker),
            )
            for existing_field, existing_marker in markers_by_record[record_number]
        }
        if marker and marker_key not in existing_keys:
            markers_by_record[record_number].append(field_marker)
    return dict(markers_by_record)


def apply_markers_to_tag_values(
    tag_values: Iterable[tuple[int, str]],
    field_markers: Iterable[tuple[int, str]],
    *,
    age_marker: str = DEFAULT_AGE_MARKER,
    age_marker_field: int = DEFAULT_AGE_MARKER_FIELD,
    stats: MarkerApplicationStats | None = None,
) -> tuple[list[tuple[int, str]], bool]:
    """Добавляет метки прямо в набор полей записи ИРБИС без дублей.

    Если более старая версия программы уже успела записать одну и ту же
    длинную метку дважды (двумя повторениями поля или два раза внутри одного
    поля), при следующей обработке записи оставляется ровно один экземпляр.
    """
    fields = [(int(tag), str(value)) for tag, value in tag_values]

    # Дедуплицируем запросы не только побайтово, но и по отображаемому тексту:
    # NBSP/обычный пробел и другие совместимые Unicode-варианты не должны
    # порождать два одинаковых уведомления.
    requested: list[tuple[int, str]] = []
    requested_keys: set[tuple[int, str]] = set()
    for tag, marker in field_markers:
        marker = str(marker).strip()
        if not marker:
            continue
        key = (int(tag), _normalized_marker_text(marker))
        if key in requested_keys:
            continue
        requested_keys.add(key)
        requested.append((int(tag), marker))

    if stats is not None:
        for tag, marker in requested:
            marker_identity = _foreign_marker_identity(marker)
            occurrences = 0
            for existing_tag, value in fields:
                if existing_tag != tag:
                    continue
                same_foreign_marker = (
                    marker_identity is not None
                    and _foreign_marker_identity(value) == marker_identity
                )
                if not same_foreign_marker and not _field_contains_marker(value, marker):
                    continue
                repeat_count = _marker_only_repeat_count(value, marker)
                occurrences += max(1, repeat_count)
            if occurrences:
                stats.already_present += 1
            else:
                stats.added += 1

    changed = False

    # Отдельно схлопываем короткие авторские метки ^AI^@... по смысловому
    # ключу. Это закрывает случай из АРМ Каталогизатора, когда два повторения
    # выглядят абсолютно одинаково, но одно содержит невидимый Unicode-символ.
    requested_foreign_keys = {
        (tag, key)
        for tag, marker in requested
        if (key := _foreign_marker_identity(marker)) is not None
    }
    if requested_foreign_keys:
        seen_foreign: set[tuple[int, str]] = set()
        deduped_fields: list[tuple[int, str]] = []
        for tag, value in fields:
            foreign_key = _foreign_marker_identity(value)
            pair = (tag, foreign_key) if foreign_key is not None else None
            if pair is not None and pair in requested_foreign_keys:
                if pair in seen_foreign:
                    changed = True
                    if stats is not None:
                        stats.duplicates_repaired += 1
                    continue
                seen_foreign.add(pair)
            deduped_fields.append((tag, value))
        fields = deduped_fields

    # Исправляем уже существующие дубли именно наших запрошенных меток.
    # Чужие/служебные значения в том же поле не трогаем.
    for tag, marker in requested:
        matching_indices: list[int] = []
        for index, (existing_tag, value) in enumerate(fields):
            if existing_tag != tag or not _field_contains_marker(value, marker):
                continue
            repeat_count = _marker_only_repeat_count(value, marker)
            if repeat_count > 1:
                fields[index] = (existing_tag, marker)
                changed = True
                if stats is not None:
                    stats.duplicates_repaired += repeat_count - 1
            matching_indices.append(index)

        # Если одинаковая метка записана отдельными повторениями одного поля,
        # оставляем первое. Удаляем только marker-only повторения, чтобы не
        # потерять дополнительное содержимое служебного поля.
        if len(matching_indices) > 1:
            keep_index = matching_indices[0]
            for index in reversed(matching_indices[1:]):
                if _marker_only_repeat_count(fields[index][1], marker) == 1:
                    del fields[index]
                    changed = True
                    if stats is not None:
                        stats.duplicates_repaired += 1

    age_marker = age_marker.strip()
    age_positions = [index for index, (tag, _value) in enumerate(fields) if tag == int(age_marker_field)]
    if stats is not None and age_marker:
        if not age_positions:
            stats.added += 1
        else:
            for index in age_positions:
                _tag, value = fields[index]
                if _field_contains_marker(value, age_marker):
                    stats.already_present += 1
                else:
                    stats.added += 1
    if age_marker and age_positions:
        for index in age_positions:
            tag, value = fields[index]
            if not _field_contains_marker(value, age_marker):
                fields[index] = (tag, value + age_marker)
                changed = True

    existing_by_tag: dict[int, list[str]] = defaultdict(list)
    for tag, value in fields:
        existing_by_tag[tag].append(value)

    def insert_ordered(tag: int, value: str) -> None:
        position = len(fields)
        for index, (existing_tag, _existing_value) in enumerate(fields):
            if existing_tag > tag:
                position = index
                break
        fields.insert(position, (tag, value))

    for tag, marker in requested:
        requested_foreign_key = _foreign_marker_identity(marker)
        already_present = any(_field_contains_marker(value, marker) for value in existing_by_tag[tag])
        if not already_present and requested_foreign_key is not None:
            already_present = any(
                _foreign_marker_identity(value) == requested_foreign_key
                for value in existing_by_tag[tag]
            )
        if not already_present:
            insert_ordered(tag, marker)
            existing_by_tag[tag].append(marker)
            changed = True

    if age_marker and not age_positions:
        insert_ordered(int(age_marker_field), age_marker)
        changed = True

    return fields, changed


def export_modified_database(
    source_path: str | Path,
    output_path: str | Path,
    results: list[MatchResult],
    summary: ComparisonSummary,
    progress_cb: Optional[ProgressCallback] = None,
    cancel_cb: Optional[CancelCallback] = None,
    *,
    substance_marker: str = DEFAULT_SUBSTANCE_MARKER,
    foreign_agent_marker_template: str = DEFAULT_FOREIGN_AGENT_MARKER_TEMPLATE,
    age_marker: str = DEFAULT_AGE_MARKER,
    substance_marker_field: int = DEFAULT_SUBSTANCE_MARKER_FIELD,
    foreign_agent_marker_field: int = DEFAULT_FOREIGN_AGENT_MARKER_FIELD,
    age_marker_field: int = DEFAULT_AGE_MARKER_FIELD,
) -> Path:
    """Сохраняет копию TXT-базы и помечает только найденные записи."""
    source_path = Path(source_path)
    output_path = Path(output_path)
    if source_path.resolve() == output_path.resolve():
        raise ValueError("Изменённая TXT-база не должна перезаписывать исходный файл.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if progress_cb:
        progress_cb(94, "Добавление меток в найденные записи TXT")

    matched_numbers = {
        result.database.source_record_number or result.database.record_number
        for result in results
        if _result_is_eligible_for_txt_marker(result)
        and result.database is not None
        and (
            not result.database.source_file
            or Path(result.database.source_file).resolve() == source_path.resolve()
        )
    }
    markers_by_record: dict[int, list[tuple[int, str]]] = defaultdict(list)
    for result in results:
        if not _result_is_eligible_for_txt_marker(result) or result.database is None:
            continue
        record = result.database
        if record.source_file and Path(record.source_file).resolve() != source_path.resolve():
            continue
        record_number = record.source_record_number or record.record_number
        if result.source_type == SOURCE_FOREIGN_AGENTS:
            agent_name = _foreign_agent_marker_name(result)
            marker = foreign_agent_marker_template.replace("{name}", agent_name)
            marker_field = foreign_agent_marker_field
        else:
            marker = substance_marker
            marker_field = substance_marker_field
        marker = marker.strip()
        field_marker = (marker_field, marker)
        marker_key = (int(marker_field), _normalized_marker_text(marker))
        existing_keys = {
            (int(existing_field), _normalized_marker_text(existing_marker))
            for existing_field, existing_marker in markers_by_record[record_number]
        }
        if marker and marker_key not in existing_keys:
            markers_by_record[record_number].append(field_marker)
    text, encoding = _read_text_file_with_encoding(source_path)
    newline = _detect_newline(text)

    # Разделители ***** сохраняются без изменений. Чётные части — записи, нечётные — разделители.
    parts = re.split(r"(\r?\n\*{5}\s*(?:\r?\n|$))", text)
    record_number = 0
    modified_count = 0

    for part_index in range(0, len(parts), 2):
        raw_record = parts[part_index]
        if not raw_record.strip():
            continue
        record_number += 1
        if record_number not in matched_numbers:
            continue
        if modified_count % 250 == 0:
            _cancelled(cancel_cb)
        modified_record, changed = _modify_matched_record(
            raw_record,
            newline,
            [],
            age_marker.strip(),
            age_field=age_marker_field,
            field_markers=markers_by_record.get(record_number, []),
        )
        parts[part_index] = modified_record
        if changed:
            modified_count += 1

    with output_path.open("w", encoding=encoding, newline="") as output_file:
        output_file.write("".join(parts))
    summary.modified_database_file = str(output_path)
    summary.modified_database_records = len(matched_numbers)

    if progress_cb:
        progress_cb(100, f"Готово: помечено записей TXT — {len(matched_numbers):,}")
    return output_path


def _modified_output_for_source(output_path: Path, source_path: Path, multiple: bool) -> Path:
    if not multiple:
        return output_path
    return output_path.with_name(f"{output_path.stem}_{source_path.stem}{output_path.suffix}")


def export_modified_databases(
    source_paths: list[str | Path],
    output_path: str | Path,
    results: list[MatchResult],
    summary: ComparisonSummary,
    progress_cb: Optional[ProgressCallback] = None,
    cancel_cb: Optional[CancelCallback] = None,
    *,
    substance_marker: str = DEFAULT_SUBSTANCE_MARKER,
    foreign_agent_marker_template: str = DEFAULT_FOREIGN_AGENT_MARKER_TEMPLATE,
    age_marker: str = DEFAULT_AGE_MARKER,
    substance_marker_field: int = DEFAULT_SUBSTANCE_MARKER_FIELD,
    foreign_agent_marker_field: int = DEFAULT_FOREIGN_AGENT_MARKER_FIELD,
    age_marker_field: int = DEFAULT_AGE_MARKER_FIELD,
) -> list[Path]:
    output = Path(output_path)
    multiple = len(source_paths) > 1
    written: list[Path] = []
    total_marked = 0
    files = []
    for source in source_paths:
        source_path = Path(source)
        target = _modified_output_for_source(output, source_path, multiple)
        export_modified_database(
            source_path,
            target,
            results,
            summary,
            progress_cb,
            cancel_cb,
            substance_marker=substance_marker,
            foreign_agent_marker_template=foreign_agent_marker_template,
            age_marker=age_marker,
            substance_marker_field=substance_marker_field,
            foreign_agent_marker_field=foreign_agent_marker_field,
            age_marker_field=age_marker_field,
        )
        written.append(target)
        files.append(str(target))
        total_marked += sum(
            1
            for record_number in {
                result.database.source_record_number or result.database.record_number
                for result in results
                if _result_is_eligible_for_txt_marker(result)
                and result.database is not None
                and result.database.source_file
                and Path(result.database.source_file).resolve() == source_path.resolve()
            }
        )
    summary.modified_database_file = "; ".join(files)
    summary.modified_database_records = total_marked
    return written


def compare_and_export(
    database_path: str | Path | list[str | Path],
    excel_paths: list[str | Path],
    output_path: str | Path,
    modified_database_path: str | Path,
    *,
    foreign_agents_path: str | Path | None = None,
    use_isbn_matching: bool = True,
    use_title_fallback: bool = True,
    use_fuzzy: bool = False,
    fuzzy_threshold: int = 90,
    report_options: dict[str, Any] | None = None,
    substance_marker: str = DEFAULT_SUBSTANCE_MARKER,
    foreign_agent_marker_template: str = DEFAULT_FOREIGN_AGENT_MARKER_TEMPLATE,
    age_marker: str = DEFAULT_AGE_MARKER,
    substance_marker_field: int = DEFAULT_SUBSTANCE_MARKER_FIELD,
    foreign_agent_marker_field: int = DEFAULT_FOREIGN_AGENT_MARKER_FIELD,
    age_marker_field: int = DEFAULT_AGE_MARKER_FIELD,
    progress_cb: Optional[ProgressCallback] = None,
    cancel_cb: Optional[CancelCallback] = None,
) -> tuple[list[MatchResult], ComparisonSummary]:
    database_paths = database_path if isinstance(database_path, list) else [database_path]
    results, summary = compare_files(
        database_paths,
        excel_paths,
        foreign_agents_path=foreign_agents_path,
        use_isbn_matching=use_isbn_matching,
        use_title_fallback=use_title_fallback,
        use_fuzzy=use_fuzzy,
        fuzzy_threshold=fuzzy_threshold,
        progress_cb=progress_cb,
        cancel_cb=cancel_cb,
    )
    if report_options is None or report_options.get("enabled", True):
        export_results(
            output_path,
            results,
            summary,
            progress_cb,
            cancel_cb,
            report_options=report_options,
        )
    if not (report_options and report_options.get("report_only", False)):
        export_modified_databases(
            database_paths,
            modified_database_path,
            results,
            summary,
            progress_cb,
            cancel_cb,
            substance_marker=substance_marker,
            foreign_agent_marker_template=foreign_agent_marker_template,
            age_marker=age_marker,
            substance_marker_field=substance_marker_field,
            foreign_agent_marker_field=foreign_agent_marker_field,
            age_marker_field=age_marker_field,
        )
    return results, summary


def result_to_dict(result: MatchResult) -> dict[str, Any]:
    data = asdict(result)
    return data
