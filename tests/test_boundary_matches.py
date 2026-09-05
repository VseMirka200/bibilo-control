import tempfile
import unittest
from dataclasses import replace
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from irbis_control.core.matcher import (
    SOURCE_FOREIGN_AGENTS,
    SOURCE_SUBSTANCES,
    EXTRA_MATCH_RULES,
    parse_match_rule,
    ComparisonSummary,
    DatabaseIndex,
    DatabaseRecord,
    ExcelEntry,
    ForeignAgentEntry,
    MatchResult,
    build_markers_by_record,
    compare_foreign_agents,
    compare_database_records,
    compare_files,
    compare_and_export,
    database_record_from_tag_values,
    _detect_header,
    _make_entry,
    _deduplicate_cross_sheet_entries,
    _read_xlsx_entries,
    export_results,
)


class BoundaryMatchTests(unittest.TestCase):
    def test_new_rules_match_independently_and_can_be_disabled(self) -> None:
        record = database_record_from_tag_values(1, [
            (10, "^A9780306406157"), (200, "^AТестовая книга"), (700, "^AИванов^BИ.И."), (210, "^CАСТ^D2024"),
        ])
        entry = ExcelEntry(1, "books.xlsx", "Книги", 2, author="Иванов И.И.",
                           title="Тестовая книга", isbn="9780306406157", publisher="«АСТ»", year="2024 г.")
        index = DatabaseIndex([record])
        for key, (label, required) in EXTRA_MATCH_RULES.items():
            with self.subTest(rule=key):
                result = index.match(entry, False, False, False, 90, {key: True})[0]
                self.assertEqual(label, result.method)
                weak = key in {"title_year", "title_publisher", "author_year", "author_publisher", "author_publisher_year"}
                self.assertEqual("Возможное совпадение" if weak else "Совпадение", result.status)
                self.assertEqual(not weak, 1 in build_markers_by_record([result]))
                disabled = index.match(entry, False, False, False, 90, {key: False})
                self.assertEqual({}, build_markers_by_record(disabled))
                for missing in required:
                    incomplete = replace(entry, **{missing: ""})
                    result = index.match(incomplete, False, False, False, 90, {key: True})
                    self.assertEqual({}, build_markers_by_record(result))

    def test_rule_parser_normalizes_order_aliases_and_rejects_errors(self) -> None:
        for label, expected in (
            ("Год издания + Издатель + Заглавие", "title_publisher_year"),
            ("publisher, ISBN", "isbn_publisher"),
            ("Название + Автор", "use_title_fallback"),
            ("исбн", "use_isbn_matching"),
            ("год＋название", "title_year"),
            ("Автор + издательство + год", "author_publisher_year"),
        ):
            self.assertEqual(expected, parse_match_rule(label))
        for invalid in ("", "Название +", "Издательство + год", "Автор", "Название", "Название + Название", "ISBN + ИСБН", "Название + цена"):
            with self.subTest(invalid=invalid), self.assertRaises(ValueError):
                parse_match_rule(invalid)
        for key, (label, _fields) in EXTRA_MATCH_RULES.items():
            self.assertEqual(key, parse_match_rule(label))

    def test_isbn_combination_works_without_title_and_requires_metadata(self) -> None:
        record = DatabaseRecord(1, isbns=["9780306406157"], publication=["^CАСТ^D2024"])
        entry = ExcelEntry(1, "books.xlsx", "Книги", 2, isbn="9780306406157", year="2024")
        index = DatabaseIndex([record])
        result = index.match(entry, False, False, False, 90, {"isbn_year": True})[0]
        self.assertEqual("ISBN + год", result.method)
        self.assertEqual("Совпадение", result.status)
        for changed in (replace(entry, year="2023"), replace(entry, isbn="9780306406158")):
            self.assertEqual("Не найдено", index.match(changed, False, False, False, 90, {"isbn_year": True})[0].status)

    def test_placeholder_publisher_never_confirms_a_match(self) -> None:
        for publisher in ("нет", "[б. и.]", "не указано", "n/a"):
            entry = ExcelEntry(1, "books.xlsx", "Книги", 2, title="Книга", publisher=publisher, year="2024")
            record = DatabaseRecord(1, titles=[entry.title], publication=[f"^C{publisher}^D2024"])
            results = DatabaseIndex([record]).match(entry, False, False, False, 90, {"title_publisher_year": True})
            self.assertEqual({}, build_markers_by_record(results))

    def test_author_publication_rule_without_title_or_isbn_requires_review(self) -> None:
        records = [
            DatabaseRecord(1, authors=["Иванов И.И."], publication=["^CАСТ^D2024"]),
            DatabaseRecord(2, authors=["Петров П.П."], publication=["^CАСТ^D2024"]),
            DatabaseRecord(3, authors=["Иванов И.И."], publication=["^CЭксмо^D2024"]),
            DatabaseRecord(4, authors=["Иванов И.И."], publication=["^CАСТ^D2023"]),
        ]
        entry = ExcelEntry(1, "books.xlsx", "Книги", 2, author="Иванов И.И.", publisher="АСТ", year="2024")
        rules = {"author_publisher_year": True}
        results = DatabaseIndex(records).match(entry, False, False, False, 90, rules)
        self.assertEqual([1], [result.database.record_number for result in results])
        self.assertEqual("Возможное совпадение", results[0].status)
        self.assertEqual({}, build_markers_by_record(results))
        with patch("irbis_control.core.matcher.read_excel_entries", return_value=([entry], [])), patch(
            "irbis_control.core.matcher.read_foreign_agent_entries", return_value=([], [])
        ):
            _, summary = compare_database_records(records, ["books.xlsx"], use_isbn_matching=False,
                                                  use_title_fallback=False, match_rules=rules)
        self.assertEqual(1, summary.review_rows)
        self.assertEqual(0, summary.matched_excel_rows)

    def test_confirmed_rule_wins_over_author_review(self) -> None:
        record = DatabaseRecord(1, titles=["Книга"], authors=["Иванов И.И."], publication=["^CАСТ^D2024"])
        entry = ExcelEntry(1, "books.xlsx", "Книги", 2, title="Книга", author="Иванов И.И.", publisher="АСТ", year="2024")
        results = DatabaseIndex([record]).match(entry, False, False, False, 90,
                                               {"author_publisher_year": True, "title_publisher_year": True})
        self.assertEqual("Название + издательство + год", results[0].method)
        self.assertIn(1, build_markers_by_record(results))

    def test_strong_rule_wins_over_review_and_no_duplicate_record(self) -> None:
        record = DatabaseRecord(1, titles=["Тестовая книга", "Тестовая книга"], authors=["Иванов И.И."], publication=["^CАСТ^D2024"])
        entry = ExcelEntry(1, "books.xlsx", "Книги", 2, title=record.main_title, author="Иванов И.И.", publisher="АСТ", year="2024")
        rules = {"title_year": True, "title_author_publisher_year": True}
        results = DatabaseIndex([record]).match(entry, False, False, False, 90, rules)
        self.assertEqual(1, len(results))
        self.assertEqual("Название + автор + издательство + год", results[0].method)
        self.assertEqual(100.0, results[0].confidence)
        # An enabled legacy author rule must not be weakened by a review-only rule.
        result = DatabaseIndex([record]).match(entry, False, True, False, 90, {"title_year": True})[0]
        self.assertEqual("Название и автор", result.method)
        self.assertEqual("Совпадение", result.status)

    def test_publisher_and_year_distinguish_editions_without_author(self) -> None:
        records = [
            DatabaseRecord(1, titles=["Тестовая книга"], publication=["^CАСТ^D2024"]),
            DatabaseRecord(2, titles=["Тестовая книга"], publication=["^CАСТ^D2023"]),
            DatabaseRecord(3, titles=["Тестовая книга"], publication=["^CЭксмо^D2024"]),
            DatabaseRecord(4, titles=["Тестовая книга"], publication=["^CАСТ", "^D2024"]),
        ]
        entry = ExcelEntry(1, "books.xlsx", "Книги", 2, title="Тестовая книга", publisher="АСТ", year="2024")
        results = DatabaseIndex(records).match(entry, False, False, False, 90, {"title_publisher_year": True})
        self.assertEqual([1], [result.database.record_number for result in results])

    def test_new_rules_reject_missing_database_fields_and_conflicting_author(self) -> None:
        entry = ExcelEntry(1, "books.xlsx", "Книги", 2, title="Тестовая книга", author="Иванов И.И.",
                           publisher="АСТ", year="2024")
        for publication, authors in [("^CАСТ", []), ("^D2024", []), ("^CАСТ^D2024", ["Петров П.П."])]:
            with self.subTest(publication=publication, authors=authors):
                record = DatabaseRecord(1, titles=[entry.title], authors=authors, publication=[publication])
                results = DatabaseIndex([record]).match(entry, False, False, False, 90, {"title_publisher_year": True})
                self.assertEqual({}, build_markers_by_record(results))

    def test_new_rules_do_not_treat_ambiguous_year_as_exact(self) -> None:
        record = DatabaseRecord(1, titles=["Тестовая книга"], publication=["^CАСТ^D2024"])
        for year in ("2023–2024", "2024?", "", "нет"):
            entry = ExcelEntry(1, "books.xlsx", "Книги", 2, title=record.main_title, publisher="АСТ", year=year)
            results = DatabaseIndex([record]).match(entry, False, False, False, 90, {"title_publisher_year": True})
            self.assertEqual({}, build_markers_by_record(results))

    def test_excel_publisher_and_year_headers_and_cross_sheet_editions(self) -> None:
        rows = [("Название", "Издательство", "Год издания"), ("Книга", "АСТ", 2024.0)]
        _, mapping, headers = _detect_header(rows)
        entry = _make_entry(1, Path("books.xlsx"), "Книги", 2, rows[1], mapping, headers)
        self.assertEqual(("АСТ", "2024"), (entry.publisher, entry.year))
        other_year = replace(entry, entry_id=2, sheet_name="Лист2", year="2023")
        other_publisher = replace(entry, entry_id=3, sheet_name="Лист2", publisher="Эксмо")
        duplicate = replace(entry, entry_id=4, sheet_name="Лист2")
        entries, skipped = _deduplicate_cross_sheet_entries([entry, other_year, other_publisher, duplicate])
        self.assertEqual([1, 2, 3], [item.entry_id for item in entries])
        self.assertEqual(1, skipped)
        with self.assertRaises(ValueError):
            _detect_header([("Издательство", "Год")])

    def test_rules_reach_direct_and_txt_comparison_and_summary(self) -> None:
        record = DatabaseRecord(1, titles=["Книга"], publication=["^CАСТ^D2024"])
        entry = ExcelEntry(1, "books.xlsx", "Книги", 2, title="Книга", publisher="АСТ", year="2024")
        options = dict(use_isbn_matching=False, use_title_fallback=False, match_rules={"title_publisher_year": True})
        for with_foreign in (False, True):
            foreign = [self._foreign_entry(1, "Петров Петр Петрович")] if with_foreign else []
            with patch("irbis_control.core.matcher.read_excel_entries", return_value=([entry], [])), patch(
                "irbis_control.core.matcher.read_foreign_agent_entries", return_value=(foreign, [])
            ), patch("irbis_control.core.matcher.parse_database", return_value=[record]):
                for results, summary in (
                    compare_database_records([record], ["books.xlsx"], **options),
                    compare_files(["database.txt"], ["books.xlsx"], **options),
                ):
                    self.assertEqual(1, summary.matched_excel_rows)
                    self.assertEqual(1, summary.exact_title_rows)
                    self.assertIn(1, build_markers_by_record(results))

    def test_xlsx_reader_loads_publication_fields(self) -> None:
        workbook = Workbook()
        workbook.active.append(["Title", "Publisher", "Year"])
        workbook.active.append(["Книга", "АСТ", 2024])
        with patch("irbis_control.core.matcher._load_workbook_quiet", return_value=workbook):
            entries, warnings = _read_xlsx_entries(Path("books.xlsx"), 1)
        self.assertFalse(warnings)
        self.assertEqual(("АСТ", "2024"), (entries[0].publisher, entries[0].year))

    def test_new_rule_reaches_report_only_export_without_writing_files(self) -> None:
        record = DatabaseRecord(1, titles=["Тестовая книга"], publication=["^CАСТ^D2024"])
        entry = ExcelEntry(1, "books.xlsx", "Книги", 2, title=record.main_title, publisher="АСТ", year="2024")
        output = BytesIO()
        with patch("irbis_control.core.matcher.read_excel_entries", return_value=([entry], [])), patch(
            "irbis_control.core.matcher.read_foreign_agent_entries", return_value=([], [])
        ), patch("irbis_control.core.matcher.parse_database", return_value=[record]), patch(
            "irbis_control.core.matcher.atomic_write_via_path", side_effect=lambda _path, writer: writer(output)
        ), patch("irbis_control.core.matcher.Path.mkdir"):
            results, summary = compare_and_export(
                "database.txt", ["books.xlsx"], "report.xlsx", "modified.txt",
                use_isbn_matching=False, use_title_fallback=False,
                match_rules={"title_publisher_year": True},
                report_options={"enabled": True, "substances": True, "report_only": True},
            )
        output.seek(0)
        workbook = load_workbook(output, read_only=True)
        try:
            cells = [value for sheet in workbook for row in sheet.iter_rows(values_only=True) for value in row]
            self.assertTrue(any(
                isinstance(value, str) and value.startswith(EXTRA_MATCH_RULES["title_publisher_year"][0])
                for value in cells
            ))
            self.assertEqual(1, summary.matched_excel_rows)
            self.assertEqual(0, summary.modified_database_records)
        finally:
            workbook.close()

    @staticmethod
    def _foreign_entry(entry_id: int, name: str) -> ForeignAgentEntry:
        return ForeignAgentEntry(
            entry_id=entry_id,
            source_file="foreign.xlsx",
            sheet_name="Реестр",
            row_number=entry_id + 1,
            registry_number=str(entry_id),
            name=name,
            agent_type="Физическое лицо",
        )

    def test_single_initial_with_multiple_people_requires_review(self) -> None:
        record = DatabaseRecord(record_number=10, authors=["Иванов И."])
        entries = [
            self._foreign_entry(1, "Иванов Иван Петрович"),
            self._foreign_entry(2, "Иванов Илья Сергеевич"),
        ]

        results = compare_foreign_agents([record], entries)

        self.assertEqual(2, len(results))
        self.assertTrue(all(result.status == "Возможное совпадение" for result in results))
        self.assertTrue(all(result.confidence == 90.0 for result in results))
        self.assertEqual({}, build_markers_by_record(results))

    def test_single_initial_is_not_automatic_even_for_one_candidate(self) -> None:
        record = DatabaseRecord(record_number=11, authors=["Иванов И."])

        results = compare_foreign_agents(
            [record], [self._foreign_entry(1, "Иванов Иван Петрович")]
        )

        self.assertEqual("Возможное совпадение", results[0].status)
        self.assertEqual({}, build_markers_by_record(results))

    def test_exact_title_without_author_requires_review(self) -> None:
        record = DatabaseRecord(
            record_number=20,
            titles=["Принцип сперматозоида"],
            authors=["Петров П.П."],
        )
        entry = ExcelEntry(
            entry_id=1,
            source_file="substances.xlsx",
            sheet_name="Книги",
            row_number=2,
            title="Принцип сперматозоида",
        )

        results = DatabaseIndex([record]).match(entry, True, True, False, 90)

        self.assertEqual("Возможное совпадение", results[0].status)
        self.assertEqual("Только название", results[0].method)
        self.assertEqual({}, build_markers_by_record(results))

    def test_exact_title_and_author_remain_confirmed(self) -> None:
        record = DatabaseRecord(
            record_number=21,
            titles=["Принцип сперматозоида"],
            authors=["Петров П.П."],
        )
        entry = ExcelEntry(
            entry_id=1,
            source_file="substances.xlsx",
            sheet_name="Книги",
            row_number=2,
            author="Петров П.П.",
            title="Принцип сперматозоида",
        )

        results = DatabaseIndex([record]).match(entry, True, True, False, 90)

        self.assertEqual("Совпадение", results[0].status)
        self.assertEqual(100.0, results[0].confidence)
        self.assertIn(21, build_markers_by_record(results))

    def test_confidence_below_100_is_never_eligible_for_marker(self) -> None:
        result = MatchResult(
            status="Совпадение",
            method="Название и автор",
            confidence=99.0,
            excel=ExcelEntry(1, "source.xlsx", "Лист1", 2),
            database=DatabaseRecord(record_number=30),
            source_type=SOURCE_SUBSTANCES,
        )

        self.assertEqual({}, build_markers_by_record([result]))

    def test_excel_report_contains_review_sheet_with_reason_and_source(self) -> None:
        result = MatchResult(
            status="Возможное совпадение",
            method="Только название",
            confidence=75.0,
            excel=ExcelEntry(1, "source.xlsx", "Книги", 12, title="Книга"),
            database=DatabaseRecord(record_number=40, titles=["Книга"]),
            note="Точно совпало только название",
            source_type=SOURCE_SUBSTANCES,
            matched_value="Книга",
        )
        summary = ComparisonSummary(
            database_file="Тест",
            excel_files=["source.xlsx"],
            database_records=1,
            database_records_with_isbn=0,
            excel_rows=1,
            matched_excel_rows=0,
            unmatched_excel_rows=1,
            result_rows=1,
            exact_isbn_rows=0,
            exact_title_rows=0,
            probable_rows=1,
            review_rows=1,
        )

        with tempfile.TemporaryDirectory(dir=Path.cwd()) as temp_dir:
            output = Path(temp_dir) / "report.xlsx"
            export_results(
                output,
                [result],
                summary,
                report_options={"substances": True},
            )
            workbook = load_workbook(output, read_only=True)
            try:
                self.assertIn("Требует проверки", workbook.sheetnames)
                sheet = workbook["Требует проверки"]
                self.assertEqual("Точно совпало только название", sheet["D2"].value)
                self.assertEqual("source.xlsx", sheet["I2"].value)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
