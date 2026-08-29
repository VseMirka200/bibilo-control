from __future__ import annotations

import re
import unicodedata
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable


@dataclass
class ResultDiffRow:
    change_type: str
    key: str
    values: dict[str, str]
    changed_fields: list[str] = field(default_factory=list)
    previous_values: dict[str, str] = field(default_factory=dict)


@dataclass
class ResultDiffSummary:
    old_file: str
    new_file: str
    output_file: str
    added: int
    removed: int
    changed: int
    unchanged: int
    total_changes: int
    warnings: list[str] = field(default_factory=list)


@dataclass
class TextDiffRow:
    change_type: str
    record_number: int
    old_text: str = ""
    new_text: str = ""


@dataclass
class TextDiffSummary:
    old_file: str
    new_file: str
    output_file: str
    added: int
    removed: int
    changed: int
    unchanged: int
    total_changes: int


CANONICAL_HEADERS = [
    "Раздел отчёта",
    "Автор",
    "Название",
    "ISBN",
    "Инвентарные номера",
    "Издание: город, издательство, год",
    "Номер записи в текстовой базе",
]

HEADER_ALIASES = {
    "раздел отчета": "Раздел отчёта",
    "источник совпадения": "Раздел отчёта",
    "автор": "Автор",
    "авторы": "Автор",
    "название": "Название",
    "заглавие": "Название",
    "isbn": "ISBN",
    "инвентарные номера": "Инвентарные номера",
    "инвентарный номер": "Инвентарные номера",
    "инв номер": "Инвентарные номера",
    "издание город издательство год": "Издание: город, издательство, год",
    "издание": "Издание: город, издательство, год",
    "номер записи в текстовой базе": "Номер записи в текстовой базе",
    "номер записи txt": "Номер записи в текстовой базе",
    "номер записи": "Номер записи в текстовой базе",
}


def _load_workbook_quiet(path: str | Path, **kwargs):
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
            module=r"openpyxl\.styles\.stylesheet",
        )
        return load_workbook(path, **kwargs)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _text(value)).lower().replace("ё", "е")
    text = re.sub(r"[^0-9a-zа-я]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _canonical_header(value: Any) -> str:
    normalized = _normalize_header(value)
    return HEADER_ALIASES.get(normalized, _text(value))


def _normalize_isbn(value: Any) -> str:
    return re.sub(r"[^0-9X]", "", _text(value).upper())


def _normalize_general(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _text(value)).lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _iter_xlsx_rows(path: Path) -> list[tuple[list[str], list[list[Any]], str]]:
    workbook = _load_workbook_quiet(path, read_only=True, data_only=True)
    preferred = ["Вещества", "Иностранные агенты"]
    selected = [name for name in preferred if name in workbook.sheetnames]
    if not selected:
        legacy = "Совпадения из TXT"
        selected = [legacy] if legacy in workbook.sheetnames else [workbook.sheetnames[0]]

    sheets: list[tuple[list[str], list[list[Any]], str]] = []
    for sheet_name in selected:
        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = [_text(value) for value in next(iterator)]
        except StopIteration:
            headers = []
        rows = [list(row) for row in iterator]
        sheets.append((headers, rows, worksheet.title))
    workbook.close()
    return sheets


def _iter_xls_rows(path: Path) -> list[tuple[list[str], list[list[Any]], str]]:
    import xlrd

    workbook = xlrd.open_workbook(path)
    preferred = ["Вещества", "Иностранные агенты"]
    selected = [name for name in preferred if name in workbook.sheet_names()]
    if not selected:
        legacy = "Совпадения из TXT"
        selected = [legacy] if legacy in workbook.sheet_names() else [workbook.sheet_names()[0]]

    sheets: list[tuple[list[str], list[list[Any]], str]] = []
    for sheet_name in selected:
        sheet = workbook.sheet_by_name(sheet_name)
        headers = [_text(value) for value in sheet.row_values(0)] if sheet.nrows else []
        rows = [sheet.row_values(index) for index in range(1, sheet.nrows)]
        sheets.append((headers, rows, sheet.name))
    return sheets


def _read_report(path: str | Path) -> tuple[list[dict[str, str]], list[str]]:
    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(f"Файл не найден: {source}")

    sheets = _iter_xls_rows(source) if source.suffix.lower() == ".xls" else _iter_xlsx_rows(source)
    rows: list[dict[str, str]] = []
    warnings: list[str] = []

    for headers, raw_rows, sheet_name in sheets:
        if not headers:
            warnings.append(f"В файле «{source.name}», лист «{sheet_name}» нет заголовков.")
            continue

        canonical_headers = [_canonical_header(header) for header in headers]
        for raw_row in raw_rows:
            row = {
                canonical_headers[index]: _text(raw_row[index]) if index < len(raw_row) else ""
                for index in range(len(canonical_headers))
            }
            row.pop("№", None)
            row.setdefault("Раздел отчёта", sheet_name)
            if not row.get("Раздел отчёта"):
                row["Раздел отчёта"] = sheet_name
            if any(value for key, value in row.items() if key != "Раздел отчёта"):
                rows.append(row)

        missing = [
            header for header in CANONICAL_HEADERS
            if header != "Раздел отчёта" and header not in canonical_headers
        ]
        if missing:
            warnings.append(
                f"{source.name}, лист «{sheet_name}»: отсутствуют столбцы: {', '.join(missing)}. "
                "Сравнение выполнено по доступным данным."
            )

    if not rows and not any(headers for headers, _, _ in sheets):
        raise ValueError(f"В файле «{source.name}» нет данных для сравнения.")
    return rows, warnings

def _row_key(row: dict[str, str]) -> str:
    section = _normalize_general(row.get("Раздел отчёта", "")) or "общий"
    record_number = _normalize_general(row.get("Номер записи в текстовой базе", ""))
    if record_number:
        return f"section:{section}|record:{record_number}"

    isbn = _normalize_isbn(row.get("ISBN", ""))
    if isbn:
        return f"section:{section}|isbn:{isbn}"

    title = _normalize_general(row.get("Название", ""))
    author = _normalize_general(row.get("Автор", ""))
    inventory = _normalize_general(row.get("Инвентарные номера", ""))
    return f"section:{section}|text:{title}|{author}|{inventory}"

def _comparison_fields(old_rows: Iterable[dict[str, str]], new_rows: Iterable[dict[str, str]]) -> list[str]:
    available = set()
    for row in list(old_rows) + list(new_rows):
        available.update(row)
    ordered = [header for header in CANONICAL_HEADERS if header in available]
    extras = sorted(header for header in available if header not in set(ordered) and header != "№")
    return ordered + extras


def _index_rows(rows: list[dict[str, str]]) -> tuple[dict[str, dict[str, str]], list[str]]:
    indexed: dict[str, dict[str, str]] = {}
    warnings: list[str] = []
    counts: dict[str, int] = {}
    for row in rows:
        base_key = _row_key(row)
        counts[base_key] = counts.get(base_key, 0) + 1
        key = base_key if counts[base_key] == 1 else f"{base_key}#duplicate-{counts[base_key]}"
        if counts[base_key] > 1:
            warnings.append(
                "Найдены повторяющиеся строки с одинаковым ключом; они сравнены по порядку появления."
            )
        indexed[key] = row
    return indexed, list(dict.fromkeys(warnings))


def compare_result_rows(
    old_rows: list[dict[str, str]],
    new_rows: list[dict[str, str]],
) -> tuple[list[ResultDiffRow], dict[str, int], list[str]]:
    old_index, old_warnings = _index_rows(old_rows)
    new_index, new_warnings = _index_rows(new_rows)
    fields = _comparison_fields(old_rows, new_rows)

    differences: list[ResultDiffRow] = []
    added = removed = changed = unchanged = 0

    all_keys = sorted(set(old_index) | set(new_index))
    for key in all_keys:
        old_row = old_index.get(key)
        new_row = new_index.get(key)
        if old_row is None and new_row is not None:
            added += 1
            differences.append(ResultDiffRow("Добавлено", key, dict(new_row)))
            continue
        if new_row is None and old_row is not None:
            removed += 1
            differences.append(ResultDiffRow("Удалено", key, dict(old_row)))
            continue
        assert old_row is not None and new_row is not None

        changed_fields = [
            field
            for field in fields
            if _normalize_general(old_row.get(field, "")) != _normalize_general(new_row.get(field, ""))
        ]
        if changed_fields:
            changed += 1
            differences.append(
                ResultDiffRow(
                    "Изменено",
                    key,
                    dict(new_row),
                    changed_fields=changed_fields,
                    previous_values={field: old_row.get(field, "") for field in changed_fields},
                )
            )
        else:
            unchanged += 1

    order = {"Добавлено": 0, "Изменено": 1, "Удалено": 2}
    differences.sort(
        key=lambda item: (
            order.get(item.change_type, 99),
            _normalize_general(item.values.get("Автор", "")),
            _normalize_general(item.values.get("Название", "")),
            item.key,
        )
    )
    counts = {
        "added": added,
        "removed": removed,
        "changed": changed,
        "unchanged": unchanged,
    }
    return differences, counts, list(dict.fromkeys(old_warnings + new_warnings))


def _previous_text(diff: ResultDiffRow) -> str:
    if not diff.changed_fields:
        return ""
    lines = []
    for field in diff.changed_fields:
        before = diff.previous_values.get(field, "") or "(пусто)"
        after = diff.values.get(field, "") or "(пусто)"
        lines.append(f"{field}: {before} → {after}")
    return "\n".join(lines)


def export_result_differences(
    output_path: str | Path,
    differences: list[ResultDiffRow],
    summary: ResultDiffSummary,
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Изменения"
    headers = [
        "Тип изменения",
        *CANONICAL_HEADERS,
        "Изменённые поля",
        "Что было → что стало",
    ]
    worksheet.append(headers)

    for diff in differences:
        worksheet.append(
            [
                diff.change_type,
                *[diff.values.get(header, "") for header in CANONICAL_HEADERS],
                ", ".join(diff.changed_fields),
                _previous_text(diff),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    fills = {
        "Добавлено": PatternFill("solid", fgColor="E2F0D9"),
        "Удалено": PatternFill("solid", fgColor="FCE4D6"),
        "Изменено": PatternFill("solid", fgColor="FFF2CC"),
    }

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index in range(2, len(differences) + 2):
        change_type = worksheet.cell(row_index, 1).value
        fill = fills.get(change_type)
        for cell in worksheet[row_index]:
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.row_dimensions[row_index].height = 42

    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = "A2"
    last_column = get_column_letter(len(headers))
    worksheet.auto_filter.ref = f"A1:{last_column}{max(1, len(differences) + 1)}"
    widths = {
        "A": 18,
        "B": 22,
        "C": 34,
        "D": 52,
        "E": 24,
        "F": 24,
        "G": 42,
        "H": 24,
        "I": 28,
        "J": 76,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    if differences:
        table = Table(displayName="ResultChangesTable", ref=f"A1:{last_column}{len(differences) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    info = workbook.create_sheet("Сводка")
    info.append(["Параметр", "Значение"])
    info_rows = [
        ("Старый отчёт", summary.old_file),
        ("Новый отчёт", summary.new_file),
        ("Добавлено", summary.added),
        ("Удалено", summary.removed),
        ("Изменено", summary.changed),
        ("Без изменений", summary.unchanged),
        ("Всего изменений", summary.total_changes),
        ("Дата сравнения", datetime.now().strftime("%d.%m.%Y %H:%M:%S")),
    ]
    for row in info_rows:
        info.append(row)
    if summary.warnings:
        info.append(["Предупреждения", "\n".join(summary.warnings)])
    for cell in info[1]:
        cell.fill = header_fill
        cell.font = header_font
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 110
    for row in info.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.active = 0
    workbook.save(output)
    return output


def compare_result_files(
    old_path: str | Path,
    new_path: str | Path,
    output_path: str | Path,
) -> tuple[list[ResultDiffRow], ResultDiffSummary]:
    old_source = Path(old_path)
    new_source = Path(new_path)
    output = Path(output_path)

    if old_source.resolve() == new_source.resolve():
        raise ValueError("Старый и новый отчёты должны быть разными файлами.")
    if output.resolve() in {old_source.resolve(), new_source.resolve()}:
        raise ValueError("Файл изменений нельзя сохранять поверх старого или нового отчёта.")

    old_rows, old_warnings = _read_report(old_source)
    new_rows, new_warnings = _read_report(new_source)
    differences, counts, compare_warnings = compare_result_rows(old_rows, new_rows)
    warnings = list(dict.fromkeys(old_warnings + new_warnings + compare_warnings))

    summary = ResultDiffSummary(
        old_file=str(old_source),
        new_file=str(new_source),
        output_file=str(output),
        added=counts["added"],
        removed=counts["removed"],
        changed=counts["changed"],
        unchanged=counts["unchanged"],
        total_changes=len(differences),
        warnings=warnings,
    )
    export_result_differences(output, differences, summary)
    return differences, summary


def _read_plain_text(path: str | Path) -> str:
    source = Path(path)
    raw = source.read_bytes()
    if raw.startswith(b"\xef\xbb\xbf"):
        return raw.decode("utf-8-sig")
    for encoding in ("utf-8", "cp1251"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            pass
    raise ValueError(f"Не удалось определить кодировку файла {source.name}.")


def _split_text_records(text: str) -> list[str]:
    records = [part.strip() for part in re.split(r"\r?\n\*{5}\s*(?:\r?\n|$)", text) if part.strip()]
    return records if records else [line.rstrip() for line in text.splitlines()]


def _normalize_record_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKC", text).replace("\r\n", "\n").replace("\r", "\n")
    lines = [re.sub(r"\s+", " ", line).strip() for line in normalized.split("\n")]
    return "\n".join(line for line in lines if line)


def compare_text_records(old_records: list[str], new_records: list[str]) -> tuple[list[TextDiffRow], dict[str, int]]:
    differences: list[TextDiffRow] = []
    added = removed = changed = unchanged = 0
    total = max(len(old_records), len(new_records))

    for index in range(total):
        old_text = old_records[index] if index < len(old_records) else ""
        new_text = new_records[index] if index < len(new_records) else ""
        number = index + 1
        if old_text and not new_text:
            removed += 1
            differences.append(TextDiffRow("Удалено", number, old_text=old_text))
        elif new_text and not old_text:
            added += 1
            differences.append(TextDiffRow("Добавлено", number, new_text=new_text))
        elif _normalize_record_text(old_text) != _normalize_record_text(new_text):
            changed += 1
            differences.append(TextDiffRow("Изменено", number, old_text=old_text, new_text=new_text))
        else:
            unchanged += 1

    return differences, {
        "added": added,
        "removed": removed,
        "changed": changed,
        "unchanged": unchanged,
    }


def export_text_differences(
    output_path: str | Path,
    differences: list[TextDiffRow],
    summary: TextDiffSummary,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "Сравнение TXT-баз",
        f"Старый файл: {summary.old_file}",
        f"Новый файл: {summary.new_file}",
        f"Добавлено: {summary.added}",
        f"Удалено: {summary.removed}",
        f"Изменено: {summary.changed}",
        f"Без изменений: {summary.unchanged}",
        f"Всего изменений: {summary.total_changes}",
        f"Дата сравнения: {datetime.now():%d.%m.%Y %H:%M:%S}",
        "",
    ]
    for diff in differences:
        lines.extend([
            "=" * 80,
            f"{diff.change_type}. Запись {diff.record_number}",
            "-" * 80,
        ])
        if diff.old_text:
            lines.extend(["Было:", diff.old_text, ""])
        if diff.new_text:
            lines.extend(["Стало:", diff.new_text, ""])
    output.write_text("\n".join(lines), encoding="utf-8")
    return output


def compare_text_files(
    old_path: str | Path,
    new_path: str | Path,
    output_path: str | Path,
) -> tuple[list[TextDiffRow], TextDiffSummary]:
    old_source = Path(old_path)
    new_source = Path(new_path)
    output = Path(output_path)
    if old_source.resolve() == new_source.resolve():
        raise ValueError("Старый и новый TXT-файлы должны быть разными файлами.")
    if output.resolve() in {old_source.resolve(), new_source.resolve()}:
        raise ValueError("Файл изменений нельзя сохранять поверх старого или нового TXT-файла.")

    old_records = _split_text_records(_read_plain_text(old_source))
    new_records = _split_text_records(_read_plain_text(new_source))
    differences, counts = compare_text_records(old_records, new_records)
    summary = TextDiffSummary(
        old_file=str(old_source),
        new_file=str(new_source),
        output_file=str(output),
        added=counts["added"],
        removed=counts["removed"],
        changed=counts["changed"],
        unchanged=counts["unchanged"],
        total_changes=len(differences),
    )
    export_text_differences(output, differences, summary)
    return differences, summary
